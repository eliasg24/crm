# Django
import re
from django.db.models import Q, Max, Subquery, Count, F, Func, ExpressionWrapper, DurationField, IntegerField, DateTimeField
from django.db.models.functions import Coalesce, ExtractDay, TruncDate, Now, Extract, Cast
from django.contrib.auth import views as auth_views
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.http.response import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.timezone import make_aware
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, ListView, TemplateView, DetailView, DeleteView, UpdateView
from django.views.generic.base import View

# Functions
from dashboards import functions
from crm import settings

# Forms
from dashboards.forms import LeadForm

# Models
from django.contrib.auth.models import User, Group
from dashboards.models import Prospecto, Asesor, Catalogo, CatalogoModelo, Lead, CatalogoRespuestasByEtapa, Historial, HistorialVerificaciones, Retomas, VehiculosInteresLead, Evento

# Utilities
from datetime import date, datetime, timedelta
import json
from openpyxl import load_workbook

class DiffDays(Func):
    function = 'DATE_PART'
    template = "%(function)s('day', %(expressions)s)"

class CastDate(Func):
    function = 'date_trunc'
    template = "%(function)s('day', %(expressions)s)"

class LoginView(auth_views.LoginView):
    # Vista de Login

    template_name = "Login.html"
    redirect_authenticated_user = True


class LogoutView(LoginRequiredMixin, auth_views.LogoutView):
    # Vista de Logout
    pass

class CapturaView(LoginRequiredMixin, TemplateView):
    # Vista de Captura

    template_name = "Captura.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        nombre = functions.separar_nombre(user.username)

        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        now = datetime.now()
        
        leads = Lead.objects.all()
        prospectos = Prospecto.objects.all()
        historiales = Historial.objects.all()

        lista_historial = []

        """for historial in historiales:
            if historial.lead.id in lista_historial:
                pass
            else:
                if historial.operacion[0:31] == "El cliente a dado una respuesta":
                    lead = Lead.objects.get(id=historial.lead.id)
                    lead.fecha_primer_contacto = historial.fecha
                    try:
                        lead.tiempo_primer_contacto = (historial.fecha - lead.fecha_hora_asignacion_asesor).total_seconds() / 60
                        lead.save()
                        lista_historial.append(lead.id)
                        print(lead.id)
                    except:
                        pass"""

        """for lead in leads:
            lead.tiempo_cambio_de_etapa = None
            print(lead)
            lead.save()"""

        """for lead in leads:
            fecha_apertura = lead.fecha_apertura
            fecha = datetime(fecha_apertura.year, fecha_apertura.month, fecha_apertura.day)
            lead.fecha_apertura = fecha
            print(lead)
            lead.save()"""
            

        """for row in sheet.iter_rows(min_row=2):
            username = row[6].value
            activo = row[8].value
            if activo == "Inactivo" or activo == "AsesorInactivo":
                activo = False
            else:
                activo = True
            if username != "null" and username != "NULL" and username != "":
                user = User.objects.get(username=username)
                user.is_active = activo
                user.save()"""

        """for row in sheet.iter_rows(min_row=2):
            name = row[1].value
            group = row[3].value
            correo = row[5].value
            username = row[6].value
            password = row[7].value
            if username != "null" and username != "NULL" and username != "":
                user = User.objects.create(username=username, 
                                        password=password,
                                        first_name=name,
                                        email=correo)
                if_grupo = True
                if group == "AdminCompras":
                    group = "Admin Compras"
                elif group == "Anfitrion":
                    group = "Anfitrión"
                elif group == "Anfitrion/PefiladoraInactivo":
                    group = "Anfitrión"
                elif group == "AnfitrionInactivo":
                    group = "Anfitrión"
                elif group == "AsesorInactivo":
                    group = "Asesor"
                elif group == "CompradorInactivo":
                    group = "Comprador"
                elif group == "Jefe de Sala":
                    group = "Jefe de sala"
                elif group == "PeritoInactivo":
                    group = "Perito"
                elif group == "Lider CRM":
                    if_grupo = False
                
                if if_grupo:
                    grupo = Group.objects.get(name=group)
                    user.groups.add(grupo)"""

        """for lead in leads:
            lead.tiempo_cambio_de_etapa = None
            print(lead)
            if lead.etapa == "Contacto Asesor" and lead.respuesta == "No contesta / volver a llamar":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto Asesor" and lead.respuesta == "Llamada realizada":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto Asesor" and lead.respuesta == "Llamada Realizada":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto Asesor" and lead.respuesta == "Mensaje enviado a whatsapp":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto Asesor" and lead.respuesta == "Se deja mensaje de voz":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto Asesor":
                lead.etapa = "Interaccion"
            elif lead.etapa == "Contacto asesor" and lead.respuesta == "No contesta / volver a llamar":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto asesor" and lead.respuesta == "Llamada realizada":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto asesor" and lead.respuesta == "Llamada Realizada":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto asesor" and lead.respuesta == "Mensaje enviado a whatsapp":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto asesor" and lead.respuesta == "Se deja mensaje de voz":
                lead.etapa = "No contactado"
            elif lead.etapa == "Contacto asesor":
                lead.etapa = "Interaccion"
            elif lead.etapa == "Seguimiento" and lead.respuesta == "Espera de aprobacion de Credito":
                lead.etapa = "Oportunidad"
            elif lead.etapa == "Seguimiento" and lead.respuesta == "Aceptacion de documentos":
                lead.etapa = "Oportunidad"
            elif lead.etapa == "Seguimiento" and lead.respuesta == "Test drive":
                lead.etapa = "Oportunidad"
            elif lead.etapa == "Seguimiento" and lead.respuesta == "Negociacion":
                lead.etapa = "Oportunidad"
            elif lead.etapa == "Seguimiento":
                lead.etapa = "Interaccion"
            elif lead.etapa == "Venta y Entrega":
                lead.etapa = "Pedido"
            elif lead.etapa == "Venta y entrega":
                lead.etapa = "Pedido"
            if lead.respuesta == "Agenda cita":
                lead.respuesta = "Enviar información"
            elif lead.respuesta == "Enviar catalogo / Informacion":
                lead.respuesta = "Enviar información"
            elif lead.respuesta == "Hacer una nueva llamada_ Cita // fecha":
                lead.respuesta = "Llamada de seguimiento"
            elif lead.respuesta == "Interes mediano plazo / Fecha":
                lead.respuesta = "Interés mediano plazo / Fecha"
            elif lead.respuesta == "Interes largo plazo / Fecha":
                lead.respuesta = "Interés a largo plazo / Fecha"
            elif lead.respuesta == "En otra ciudad posibilidad traslado":
                lead.respuesta = "Llamada de seguimiento"
            elif lead.respuesta == "Vh excede su presupuesto":
                lead.respuesta = "En el momento no hay vh de interés"
            elif lead.respuesta == "No contesta / volver a llamar":
                lead.respuesta = "No contesta / Volver a llamar"
            elif lead.respuesta == "No esta el vh de su interes":
                lead.respuesta = "En el momento no hay vh de interés"
            elif lead.respuesta == "Llamada realizada" or lead.respuesta == "Llamada Realizada":
                lead.respuesta = "No contesta / Volver a llamar"
            elif lead.respuesta == "Mensaje enviado a whatsapp":
                lead.respuesta = "Esperando respuesta WhatsApp"
            elif lead.respuesta == "Se deja mensaje de voz":
                lead.respuesta = "Esperando respuesta WhatsApp"
            elif lead.respuesta == "Interes mediano plazo / Fecha":
                lead.respuesta = "Interés mediano plazo / Fecha"
            elif lead.respuesta == "Interes largo plazo / Fecha":
                lead.respuesta = "Interés a largo plazo / Fecha"
            elif lead.respuesta == "Esta interesado en otro vehiculo":
                lead.respuesta = "En el momento no hay vh de interés"
            elif lead.respuesta == "En el momento no hay vh de interes":
                lead.respuesta = "En el momento no hay vh de interés"
            elif lead.respuesta == "Espera de aprobacion de Credito":
                lead.respuesta = "En estudio de crédito"
            elif lead.respuesta == "Aceptacion de documentos":
                lead.respuesta = "En estudio de crédito"
            elif lead.respuesta == "Test drive":
                lead.respuesta = "Negociación"
            elif lead.respuesta == "Negociacion":
                lead.respuesta = "Negociación"
            elif lead.respuesta == "Marca/ Modelo no se encuntra disponible":
                lead.respuesta = "En el momento no hay vh de interés"
            elif lead.respuesta == "Alistamiento mecanico basico":
                lead.respuesta = "Alistamiento"
            elif lead.respuesta == "Entrega del vehiculo":
                lead.respuesta = "Verificación"
            elif lead.respuesta == "Entregado":
                lead.respuesta = "Entrega finalizada"
            elif lead.respuesta == "Separación":
                lead.respuesta = "Separación"
            try:
                lead.save()
            except:
                lead.delete()"""
        
        """for prospecto in prospectos:
            if len(prospecto.correo) < 4:
                prospecto.correo = ""
                prospecto.save()"""

        calendario_general = True
        origenes_lead = Catalogo.objects.filter(clasificacion="Origen Lead")
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False
                origenes_lead = Catalogo.objects.filter(clasificacion="Origen Lead Asesor")

        asesores = Asesor.objects.all()
        grupo = Group.objects.get(name="Asesor")

        marcas = CatalogoModelo.objects.all().values("marca").distinct()

        
        cantidad_morato = Lead.objects.filter(sala="Morato", fecha_apertura__month=now.month, fecha_apertura__year=now.year).count()

        cantidad_127 = Lead.objects.filter(sala="127", fecha_apertura__month=now.month, fecha_apertura__year=now.year).count()

        today = datetime.now()

        context["asesor_actual"] = asesor_actual
        context["calendario_general"] = calendario_general
        context["cantidad_morato"] = cantidad_morato
        context["cantidad_127"] = cantidad_127
        context["marcas"] = marcas
        context["nombre"] = nombre
        context["today"] = today
        context["origenes_lead"] = origenes_lead
        context["user"] = user
        return context

    def post(self, request):
        r = request.POST
        user = User.objects.get(username=self.request.user)
        
        print(r)
        if r.get("celular_verificar", None):
            prospecto = Prospecto.objects.get(celular=r.get("celular_verificar", None))
            print(prospecto)
            if prospecto:
                ultimo_lead = Lead.objects.filter(prospecto=prospecto).last()
                ultimo_lead = {"nombre_anfitrion": ultimo_lead.nombre_anfitrion, "fecha_apertura": ultimo_lead.fecha_apertura.date(), "respuesta": ultimo_lead.respuesta, "estado": ultimo_lead.estado, "nombre_asesor": ultimo_lead.nombre_asesor,}
                alerta = {"alerta_celular": True, "ultimo_lead": ultimo_lead}
            else:
                alerta = {"alerta_celular": False}

            return JsonResponse(alerta, safe=False)
        
        if r.get("marca", None):
            marca = r.get("marca", None)
            modelos = CatalogoModelo.objects.filter(marca=marca.title())
            print(marca.title())
            print("modelos")
            print(modelos)
            modelos = list(modelos.values())

            print(modelos)

            return JsonResponse(modelos, safe=False)
        if r.get("NombreProspecto", None):
            nombre = r.get("NombreProspecto", None)
            apellido_paterno = r.get("ApellidoPProspecto", None)
            apellido_materno = r.get("ApellidoMProspecto", None)
            celular = r.get("Celular", None)
            correo = r.get("Correo", None)
            origen_lead = r.get("OrigenLead", None)
            campania = r.get("Campania", None)
            tipo_documento = r.get("TipoDocumento", None)
            documento = r.get("Documento", None)
            politica_privacidad = r.get("PoliticaPrivacidad", None)
            if not politica_privacidad:
                politica_privacidad = False
            else:
                politica_privacidad = True
            anfitrion = r.get("Anfitrion", None)
            sala = r.get("Sala", None)
            nombre_asesor = r.get("Asesor", None)
            marcas_interes = r.getlist("MarcasInteres[]", None)
            modelo = json.loads(r.get("Modelo", None))
            color = json.loads(r.get("Color", None))
            marca_comentario = json.loads(r.get("MarcaComentario", None))
            print(marcas_interes)
            print(modelo)
            print(color)
            print(marca_comentario)
            m_lista = []
            for ma in range(len(marcas_interes)):
                for mo in range(len(modelo[ma])):
                    print("aver la marca de interees")
                    print(marcas_interes[ma])
                    print(modelo[ma][mo])
                    print(color[ma][mo])
                    print(marca_comentario[ma][mo])
                    m_lista.append({"marca": marcas_interes[ma],
                        "modelo": modelo[ma][mo],
                        "color": color[ma][mo],
                        "marca_comentario": marca_comentario[ma][mo],
                        "codigo": None,
                        "precio": None
                        })
            
            marcas = {"marcas": m_lista}
            comentario = r.get("Comentario", None)
            test_drive = r.get("TestDrive", None)
            if not test_drive:
                test_drive = False
            else:
                test_drive = True
            prospecto = Prospecto.objects.create(nombre=nombre,
                                apellido_paterno=apellido_paterno,
                                apellido_materno=apellido_materno,
                                celular=celular,
                                correo=correo,
                                fecha_captura=make_aware(datetime.now()),
                                nombre_asesor=nombre_asesor,
                                anfitrion=anfitrion,
                                fecha_hora_asignacion_asesor=make_aware(datetime.now()),
                                politica_privacidad=politica_privacidad,
                                )
            
            lead = Lead.objects.create(prospecto=prospecto,
                                origen_lead=origen_lead,
                                marcas_interes=marcas,
                                sala=sala,
                                etapa="No contactado",
                                respuesta="Sin contactar",
                                estado="No contactado",
                                status="Frío",
                                interes="Venta",
                                activo=True,
                                fecha_apertura=make_aware(datetime.now()),
                                comentario=comentario,
                                campania=campania,
                                tipo_documento=tipo_documento,
                                documento=documento,
                                test_drive=test_drive,
                                nombre_asesor=nombre_asesor,
                                fecha_hora_asignacion_asesor=make_aware(datetime.now()),
                                nombre_anfitrion=anfitrion,
                                )
            Historial.objects.create(lead=lead,
                        fecha=make_aware(datetime.now()),
                        responsable=user,
                        operacion=f"Creación Lead",
                        comentarios=comentario
                        )
            if marcas_interes:
                VehiculosInteresLead.objects.create(lead=lead,
                                                marca=marcas_interes[0],
                                                modelo=modelo[0][0],
                                                color=color[0][0],
                                                comentario=marca_comentario[0][0],
                                                peritaje=False,
                                                cotizar=False,
                                                aprobacion=False,
                                                precio=False,
                                                separado=False,
                                                facturado=False,
                                                mostrado=True
                                                )
            return HttpResponseRedirect(reverse_lazy('dashboards:captura'))



        if r.get("sala", None):
            sala = r.get("sala", None)
            asesores = Asesor.objects.filter(sala=sala)
            print("asesores")
            print(asesores)
            asesores = list(asesores.values())

            return JsonResponse(asesores, safe=False)

        if r.get("etapa", None):
            etapa = r.get("etapa", None)
            respuestas = CatalogoRespuestasByEtapa.objects.filter(etapa=etapa).values("respuesta").distinct()
            print("respuestas")
            print(respuestas)
            respuestas = list(respuestas)

            return JsonResponse(respuestas, safe=False)

        

class CapturaReasignamientoView(LoginRequiredMixin, DetailView):
    # Vista de Captura Reasignamiento

    template_name = "CapturaReasignamiento.html"
    slug_field = "lead"
    slug_url_kwarg = "lead"
    queryset = Lead.objects.all()
    context_object_name = "lead"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead = self.get_object()
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        asesores = Asesor.objects.all()
        grupo = Group.objects.get(name="Asesor")

        marcas = CatalogoModelo.objects.all().values("marca").distinct()

        now = datetime.now()

        cantidad_morato = Lead.objects.filter(sala="Morato", fecha_apertura__month=now.month).count()
        cantidad_127 = Lead.objects.filter(sala="127", fecha_apertura__month=now.month).count()

        today = datetime.now()

        context["asesor_actual"] = asesor_actual
        context["calendario_general"] = calendario_general
        context["cantidad_morato"] = cantidad_morato
        context["cantidad_127"] = cantidad_127
        context["marcas"] = marcas
        context["today"] = today
        context["user"] = user
        return context
    
    def post(self, request, pk):
        r = request.POST
        user = User.objects.get(username=self.request.user)
        
        print(r)
        

        if r.get("sala", None):
            sala = r.get("sala", None)
            asesores = Asesor.objects.filter(sala=sala)
            print("asesores")
            print(asesores)
            asesores = list(asesores.values())

            return JsonResponse(asesores, safe=False)

        if r.get("Asesor", None):
            sala = r.get("Sala", None)
            nombre_asesor = r.get("Asesor", None)

            lead = Lead.objects.get(pk=pk)
            lead.sala = sala
            lead.nombre_asesor = nombre_asesor
            lead.fecha_hora_asignacion_asesor = datetime.now()
            lead.fecha_hora_reasignacion = datetime.now()
            lead.tiempo_primer_contacto = None
            lead.fecha_primer_contacto = None
            lead.etapa = "No contactado"
            lead.respuesta = "Sin contactar"
            lead.estado = "No contactado"
            lead.save()
            
            print(lead)
            print(sala)
            print(nombre_asesor)

            Historial.objects.create(lead=lead,
                        fecha=datetime.now(),
                        responsable=user,
                        operacion=f"Reasignación de asesor a {nombre_asesor}",
                        comentarios=None
                        )
            return HttpResponseRedirect(reverse_lazy('dashboards:captura'))


class DetalleClienteView(LoginRequiredMixin, TemplateView):
    # Vista de Detalle Cliente

    template_name = "DetalleCliente.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        context["asesor_actual"] = asesor_actual
        context["calendario_general"] = calendario_general
        context["user"] = user

        return context

class DetalleClienteNuevoView(LoginRequiredMixin, DetailView):
    # Vista de Detalle Cliente Nuevo

    template_name = "DetalleClienteNuevo.html"
    slug_field = "lead"
    slug_url_kwarg = "lead"
    queryset = Lead.objects.all()
    context_object_name = "lead"
    form_class = LeadForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead = self.get_object()

        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        historial = Historial.objects.filter(lead=lead)

        prospecto = Prospecto.objects.get(pk=lead.prospecto.pk)
        try:
            sala = Asesor.objects.get(nombre=lead.nombre_asesor).sala
        except:
            sala = None
        asesores = Asesor.objects.filter(sala=sala)

        if lead.fecha_hora_asignacion_asesor:

            dias_totales = (datetime.now(timezone.utc) - lead.fecha_hora_asignacion_asesor).days

            tiempo_diferencia = int((datetime.now(timezone.utc) - lead.fecha_hora_asignacion_asesor).total_seconds() / 60)
        
            print(datetime.now())
            print(lead.fecha_hora_asignacion_asesor)
            print((datetime.now(timezone.utc) - lead.fecha_hora_asignacion_asesor).total_seconds())
            print(tiempo_diferencia)
            print(lead.tiempo_primer_contacto)

            if lead.tiempo_primer_contacto or tiempo_diferencia:
                functions.verificar_primer_contacto(lead, prospecto, tiempo_diferencia)

        else:
            dias_totales = 0
            tiempo_diferencia = 0

        if lead.marcas_interes:
        
            marcas_interes = eval(lead.marcas_interes)
        else:
            marcas_interes = ""
        marcas = CatalogoModelo.objects.all().values("marca").distinct()
        etapas = CatalogoRespuestasByEtapa.objects.values("etapa").distinct()
        respuestas = CatalogoRespuestasByEtapa.objects.filter(etapa=lead.etapa).values("respuesta").distinct()

        mostrar_evento = False
        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                mostrar_evento = True
                calendario_general = False
            if grupo.name == "Admin":
                mostrar_evento = True

        context["asesor_actual"] = asesor_actual
        context["asesores"] = asesores
        context["calendario_general"] = calendario_general
        context["dias_totales"] = dias_totales
        context["etapas"] = etapas
        context["historial"] = historial
        context["marcas"] = marcas
        context["marcas_interes"] = marcas_interes
        context["mostrar_evento"] = mostrar_evento
        context["prospecto"] = prospecto
        context["respuestas"] = respuestas
        context["sala"] = sala
        context["tiempo_diferencia"] = tiempo_diferencia
        context["user"] = user

        return context
    
    def post(self, request, pk):

        lead = Lead.objects.get(pk=pk)
        prospecto = Prospecto.objects.get(pk=lead.prospecto.pk)
        user = User.objects.get(username=self.request.user)
        if request.POST.get("Lead_Etapa"):
            lista_historial = []
            if lead.etapa != request.POST.get("Lead_Etapa"):
                lista_historial.append(f"Etapa. (Antes: {lead.etapa}. Ahora: {request.POST.get('Lead_Etapa')})")
            lead.etapa = request.POST.get("Lead_Etapa")
            lead.estado = request.POST.get("Lead_Etapa")
            if lead.respuesta != request.POST.get("Lead_Respuesta"):
                lista_historial.append(f"Respuesta. (Antes: {lead.respuesta}. Ahora: {request.POST.get('Lead_Respuesta')})")
            lead.respuesta = request.POST.get("Lead_Respuesta")
            comentario = request.POST.get("Lead_Comentario")
            if lead.comentario != comentario and comentario != "None":
                lista_historial.append(f"Comentario. (Antes: {lead.comentario}. Ahora: {request.POST.get('Lead_Comentario')})")
            lead.comentario = request.POST.get("Lead_Comentario")
            if request.POST.get("Lead_Respuesta") != "Sin contactar":
                if not(lead.fecha_primer_contacto):
                    lead.fecha_primer_contacto = datetime.now()
                    tiempo_primer_contacto = datetime.now() - lead.fecha_hora_asignacion_asesor.replace(tzinfo=None)
                    lead.tiempo_primer_contacto = tiempo_primer_contacto.total_seconds() / 60
                else:
                    lead.fecha_cambio_de_etapa = datetime.now()
                    tiempo_cambio_de_etapa = datetime.now() - lead.fecha_primer_contacto.replace(tzinfo=None)
                    lead.tiempo_cambio_de_etapa = tiempo_cambio_de_etapa.total_seconds() / 60
            
            marcas_interes = request.POST.getlist("MarcasInteres[]", None)
            modelo = json.loads(request.POST.get("Modelo", None))
            color = json.loads(request.POST.get("Color", None))
            codigo = json.loads(request.POST.get("Codigo", None))
            precio = json.loads(request.POST.get("Precio", None))
            
            codigo_vehiculo = request.POST.getlist("CodigoVehiculo", None)
            precio_vehiculo = request.POST.getlist("PrecioVehiculo", None)

            m_lista = eval(lead.marcas_interes)["marcas"]
            print("aver")
            print(request.POST)
            print("marcas_interes")
            print(marcas_interes)
            print("modelo")
            print(modelo)
            print("color")
            print(color)

            if len(m_lista) == 0:
                VehiculosInteresLead.objects.create(lead=lead,
                                            marca=marcas_interes[0],
                                            modelo=modelo[0][0],
                                            color=color[0][0],
                                            codigo_vehiculo=codigo[0][0],
                                            precio=precio[0][0],
                                            peritaje=False,
                                            cotizar=False,
                                            aprobacion=False,
                                            separado=False,
                                            facturado=False,
                                            mostrado=True
                                            )
            
            for ma in range(len(marcas_interes)):
                for mo in range(len(modelo[ma])):
                    m_lista.append({"marca": marcas_interes[ma],
                        "modelo": modelo[ma][mo],
                        "color": color[ma][mo],
                        "marca_comentario": None,
                        "codigo": codigo[ma][mo],
                        "precio": precio[ma][mo]
                        })
            
            marcas = {"marcas": m_lista}

            print(lead.marcas_interes)
            print(marcas)

            if str(lead.marcas_interes) != str(marcas):
                lista_historial.append(f"Marcas Interes. (Se agregaron: {m_lista})")
            lead.marcas_interes = marcas
            lead.save()
            
            if lista_historial:
                Historial.objects.create(lead=lead,
                                        fecha=datetime.now(),
                                        responsable=user,
                                        operacion=f"Se modificaron los siguientes campos: {lista_historial}",
                                        comentarios=comentario
                                        )
            return redirect("dashboards:detallenuevo", pk)
        elif request.POST.get("ProspectoNombre"):
            lista_historial = []
            if prospecto.nombre != request.POST.get("ProspectoNombre"):
                lista_historial.append("Nombre")
            prospecto.nombre = request.POST.get("ProspectoNombre")
            if prospecto.apellido_paterno != request.POST.get("ProspectoApPaterno"):
                lista_historial.append("Apellido Paterno")
            prospecto.apellido_paterno = request.POST.get("ProspectoApPaterno")
            if prospecto.apellido_materno != request.POST.get("ProspectoApMaterno"):
                lista_historial.append("Apellido Materno")
            prospecto.apellido_materno = request.POST.get("ProspectoApMaterno")
            if prospecto.celular != request.POST.get("ProspectoCelular"):
                lista_historial.append("Celular")
            prospecto.celular = request.POST.get("ProspectoCelular")
            if prospecto.correo != request.POST.get("ProspectoCorreo"):
                lista_historial.append("Correo")
            prospecto.correo = request.POST.get("ProspectoCorreo")
            prospecto.anfitrion = request.POST.get("Lead_NombreAnfitrion")
            if prospecto.contacto_nombre != request.POST.get("NombreContacto"):
                lista_historial.append("Nombre Contacto")
            prospecto.contacto_nombre = request.POST.get("NombreContacto")
            if prospecto.contacto_telefono != request.POST.get("CelularContacto"):
                lista_historial.append("Celular Contacto")
            prospecto.contacto_telefono = request.POST.get("CelularContacto")
            if request.POST.get("PoliticaDatos") == "true":
                politica_privacidad = True
            else:
                politica_privacidad = False
            if prospecto.politica_privacidad != politica_privacidad:
                lista_historial.append("Politica Datos")
            prospecto.politica_privacidad = politica_privacidad
            prospecto.save()

            if lead.nombre_anfitrion != request.POST.get("Lead_NombreAnfitrion"):
                lista_historial.append("Anfitrión")
            lead.nombre_anfitrion = request.POST.get("Lead_NombreAnfitrion")
            if lead.campania != request.POST.get("LeadCampania"):
                lista_historial.append("Campaña")
            lead.campania = request.POST.get("LeadCampania")
            if lead.tipo_documento != request.POST.get("LeadTipoDocumento"):
                lista_historial.append("Tipo de documento")
            lead.tipo_documento = request.POST.get("LeadTipoDocumento")
            if lead.documento != request.POST.get("LeadDocumento"):
                lista_historial.append("Documento")
            lead.documento = request.POST.get("LeadDocumento")
            if request.POST.get("TestDrive") == "true":
                test_drive = True
            else:
                test_drive = False
            if lead.test_drive != test_drive:
                lista_historial.append("TestDrive")
            lead.test_drive = test_drive
            lead.save()
            if lista_historial:
                Historial.objects.create(lead=lead,
                                        fecha=datetime.now(),
                                        responsable=user,
                                        operacion=f"Se modificaron los siguientes campos: {lista_historial}"
                                        )
            return JsonResponse(prospecto.pk, safe=False)
        elif request.POST.get("EstadoLlamada"):
            lista_historial = []

            estado_llamada = request.POST.get("EstadoLlamada")
            tipo_solicitud = request.POST.get("TipoSolicitud")
            reasignado = request.POST.get("AsesorReasignado")
            observaciones = request.POST.get("Observaciones")
            
            if reasignado:
                lead.nombre_asesor = reasignado
                lead.fecha_hora_reasignacion = datetime.now()
            lead.estado_llamada_verificacion = estado_llamada
            lead.tipo_solicitud_verificacion = tipo_solicitud
            lead.save()

            HistorialVerificaciones.objects.create(lead=lead,
                                                   estado_llamada=estado_llamada,
                                                   tipo_solicitud=tipo_solicitud,
                                                   responsable=user,
                                                   reasignado=reasignado,
                                                   observaciones=observaciones,
                                                   fecha_hora_verificacion=datetime.now(),
                                                   tipo="venta"
                                                   )
            if not(reasignado):
                reasignado = ""
            else:
                reasignado = f"Reasignado a: {reasignado}"


            Historial.objects.create(lead=lead,
                                    fecha=datetime.now(),
                                    responsable=user,
                                    operacion=f"Se hizo la verificación. Estado de llamada: {estado_llamada}. Tipo de solicitud: {tipo_solicitud}. {reasignado}",
                                    comentarios=observaciones
                                    )
            return JsonResponse(prospecto.pk, safe=False)
        elif request.POST.get("modelo_retoma"):
            lista_historial = []

            modelo_retoma = request.POST.get("modelo_retoma")
            valor_retoma = request.POST.get("valor_retoma")
            total_retoma = request.POST.get("total_retoma")
            total_restante = request.POST.get("total_restante")
            
            Retomas.objects.create(lead=lead,
                                        modelo=modelo_retoma,
                                        valor=valor_retoma,
                                        total=total_retoma,
                                        total_restante=total_restante,
                                        )

            Historial.objects.create(lead=lead,
                                    fecha=datetime.now(),
                                    responsable=user,
                                    operacion=f"Se hizo la retoma. Modelo: {modelo_retoma}. Valor: {valor_retoma}. Total: {total_retoma}. Total restante: {total_restante}.",
                                    )
            return JsonResponse(prospecto.pk, safe=False)
        elif request.POST.get("accion"):
            lista_historial = []

            accion = request.POST.get("accion")
            vehiculo_acciones = request.POST.get("vehiculo_acciones")
            if accion == "Separar":
                nombre_accion = "Separado"
                separado = True
                facturado = False
                mostrado = False
            elif accion == "Facturar":
                nombre_accion = "Facturado"
                separado = False
                facturado = True
                mostrado = False
            elif accion == "Mostrar":
                nombre_accion = "Mostrado en tabla"
                separado = False
                facturado = False
                mostrado = True

            marcas_interes = eval(lead.marcas_interes)

            vehiculo_acciones = vehiculo_acciones.split("/")

            if separado == True or facturado == True:
                VehiculosInteresLead.objects.filter(lead=lead, mostrado=False).delete()

            if mostrado == True:
                VehiculosInteresLead.objects.filter(lead=lead, mostrado=True).delete()
            
            try:
                precio = int(vehiculo_acciones[4])
            except:
                precio = None

            VehiculosInteresLead.objects.create(lead=lead,
                                                marca=vehiculo_acciones[0],
                                                modelo=vehiculo_acciones[1],
                                                color=vehiculo_acciones[2],
                                                codigo_vehiculo=vehiculo_acciones[3],
                                                precio=precio,
                                                peritaje=False,
                                                cotizar=False,
                                                aprobacion=False,
                                                separado=separado,
                                                facturado=facturado,
                                                mostrado=mostrado,
                                                fecha=datetime.now()
                                        )

            Historial.objects.create(lead=lead,
                                    fecha=datetime.now(),
                                    responsable=user,
                                    operacion=f"Se hizo el {nombre_accion}. Marca: {vehiculo_acciones[0]}. Modelo: {vehiculo_acciones[1]}.",
                                    )
            return JsonResponse(prospecto.pk, safe=False)
        elif request.POST.get("EventoNombre"):

            nombre = request.POST.get("EventoNombre")
            tipo = request.POST.get("EventoTipo")
            telefono_cliente = request.POST.get("EventoTelefono")
            observaciones = request.POST.get("EventoObservaciones")
            asesor = request.POST.get("EventoAsesor")
            fecha_hora = request.POST.get("EventoFechaHora")
            tiempo_evento = request.POST.get("EventoTiempo")

            print("fecha_hora")
            print(fecha_hora)
            print(type(fecha_hora))

            print(datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M"))

            print(type(datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M")))
            
            evento = Evento.objects.create(nombre=nombre,
                                           tipo=tipo,
                                           telefono_cliente=telefono_cliente,
                                           observaciones=observaciones,
                                           asesor=Asesor.objects.get(nombre=asesor),
                                           fecha_hora=datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M"),
                                           lead=lead,
                                           tiempo_evento=tiempo_evento
                                           )
            Historial.objects.create(lead=lead,
                                    fecha=datetime.now(),
                                    responsable=user,
                                    operacion=f"Se creó un evento. Nombre: {nombre}. Tipo: {tipo}. Observaciones: {observaciones}",
                                    )
            return JsonResponse(evento.pk, safe=False)
    
        if request.POST.get("select_accion"):
            accion = request.POST.get("select_accion")
            try:
                if accion == "Separar":
                    vehiculo = VehiculosInteresLead.objects.get(lead=lead, separado=True)
                elif accion == "Facturar":
                    vehiculo = VehiculosInteresLead.objects.get(lead=lead, facturado=True)
                elif accion == "Mostrar":
                    vehiculo = VehiculosInteresLead.objects.get(lead=lead, mostrado=True)
                print("vehiculo")
                print(vehiculo)
                vehiculo = f"{vehiculo.marca} {vehiculo.modelo}"
                return JsonResponse(vehiculo, safe=False)
            except:
                return JsonResponse(None, safe=False)

class OperativoAnfitrionView(LoginRequiredMixin, TemplateView):
    # Vista de Operativo Anfitrion

    template_name = "OperativoAnfitrion.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        
        leads = Lead.objects.filter(activo=True, nombre_asesor__isnull=False).order_by("-id")

        functions.verificar_primer_contacto_todos_los_leads(leads)
        
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads_agendados = Lead.objects.filter(nombre_asesor__isnull=False, activo=True).exclude(etapa="Desistido").order_by("-id")
        leads_primer_contacto = Lead.objects.filter(nombre_asesor__isnull=True, activo=True).order_by("-id")

        verificaciones = HistorialVerificaciones.objects.select_related("lead").order_by("-id")

        mostrado_marcas = VehiculosInteresLead.objects.filter(mostrado=True).values("lead").distinct().values("lead", "marca", "modelo")


        anfitriones_agendados = leads_agendados.order_by("nombre_anfitrion").values("nombre_anfitrion").distinct()
        tipos_solicitud_agendados = leads_agendados.order_by("tipo_solicitud_verificacion").values("tipo_solicitud_verificacion").distinct()
        asesores_agendados = leads_agendados.order_by("nombre_asesor").values("nombre_asesor").distinct()
        salas_agendados = leads_agendados.order_by("sala").values("sala").distinct()

        print(anfitriones_agendados)
        
        if (leads_agendados.count() % 15) == 0:
            cantidad_agendados_pag = leads_agendados.count() // 15
        else:
            cantidad_agendados_pag = leads_agendados.count() // 15 + 1
        if (leads_primer_contacto.count() % 15) == 0:
            cantidad_primer_contacto_pag = leads_primer_contacto.count() // 15
        else:
            cantidad_primer_contacto_pag = leads_primer_contacto.count() // 15 + 1

        context["anfitriones_agendados"] = anfitriones_agendados
        context["asesores_agendados"] = asesores_agendados
        context["asesor_actual"] = asesor_actual
        context["calendario_general"] = calendario_general
        context["cantidad_agendados"] = leads_agendados.count()
        context["cantidad_verificados"] = verificaciones.count()
        context["cantidad_primer_contacto"] = leads_primer_contacto.count()
        context["cantidad_agendados_pag"] = cantidad_agendados_pag
        context["cantidad_primer_contacto_pag"] = cantidad_primer_contacto_pag
        context["cantidad_verificacion_pag"] = verificaciones.count() // 15 + 1
        context["leads_agendados"] = leads_agendados[0:15]
        context["leads_primer_contacto"] = leads_primer_contacto[0:15]
        context["mostrado_marcas"] = mostrado_marcas
        context["pages_agendados"] = 1
        context["pages_primer_contacto"] = 1
        context["pages_verificacion"] = 1
        context["salas_agendados"] = salas_agendados
        context["tipos_solicitud_agendados"] = tipos_solicitud_agendados
        context["user"] = user
        context["verificaciones"] = verificaciones[0:15]

        return context
    

    def post(self, request):
        leads = Lead.objects.all()

        desde_primer_contacto = request.POST.get("desde_primer_contacto")
        hasta_primer_contacto = request.POST.get("hasta_primer_contacto")

        try:
            json.loads(request.POST.get("if_filtrar_primer_contacto"))
            if_filtrar_primer_contacto = True
        except:
            if_filtrar_primer_contacto = False
        if if_filtrar_primer_contacto:
            if desde_primer_contacto:
                desde_primer_contacto = datetime.strptime(desde_primer_contacto, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_primer_contacto)
            if hasta_primer_contacto:
                hasta_primer_contacto = datetime.strptime(hasta_primer_contacto, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_primer_contacto)

        desde_verificacion = request.POST.get("desde_verificacion")
        hasta_verificacion = request.POST.get("hasta_verificacion")

        try:
            json.loads(request.POST.get("if_filtrar_verificacion"))
            if_filtrar_verificacion = True
        except:
            if_filtrar_verificacion = False

        if if_filtrar_verificacion:
            if desde_verificacion:
                desde_verificacion = datetime.strptime(desde_verificacion, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_verificacion)
            if hasta_verificacion:
                hasta_verificacion = datetime.strptime(hasta_verificacion, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_verificacion)

        anfitrion_agendados = request.POST.get("anfitrion_agendados")
        verificado_agendados = request.POST.get("verificado_agendados")
        tipo_solicitud_agendados = request.POST.get("tipo_solicitud_agendados")
        asesor_agendados = request.POST.get("asesor_agendados")
        sala_agendados = request.POST.get("sala_agendados")
        desde_agendados = request.POST.get("desde_agendados")
        hasta_agendados = request.POST.get("hasta_agendados")
        search_agendados = request.POST.get("search_agendados")

        try:
            json.loads(request.POST.get("if_filtrar_agendados"))
            if_filtrar_agendados = True
        except:
            if_filtrar_agendados = False

        if if_filtrar_agendados:
            if anfitrion_agendados:
                leads = leads.filter(nombre_anfitrion=anfitrion_agendados)
            if verificado_agendados:
                if verificado_agendados == "SI":
                    leads = leads.filter(estado_llamada_verificacion__isnull=False)
                elif verificado_agendados == "NO":
                    leads = leads.filter(estado_llamada_verificacion__isnull=True)
            if tipo_solicitud_agendados:
                leads = leads.filter(tipo_solicitud_verificacion=tipo_solicitud_agendados)
            if asesor_agendados:
                leads = leads.filter(nombre_asesor=asesor_agendados)
            if sala_agendados:
                leads = leads.filter(sala=sala_agendados)
            if desde_agendados:
                desde_agendados = datetime.strptime(desde_agendados, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_agendados)
            if hasta_agendados:
                hasta_agendados = datetime.strptime(hasta_agendados, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_agendados)
            if search_agendados:
                leads = leads.filter(prospecto__nombre__icontains=search_agendados) | leads.filter(prospecto__celular__icontains=search_agendados)

        if request.POST.get("pages_primer_contacto"):
            page_min = (int(request.POST.get("pages_primer_contacto")) - 1) * 15
            page_max = int(request.POST.get("pages_primer_contacto")) * 15
            leads_primer_contacto = leads.filter(nombre_asesor__isnull=True, activo=True).order_by("-id")[page_min:page_max]
            leads_primer_contacto = list(leads_primer_contacto.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            return JsonResponse(leads_primer_contacto, safe=False)
        if request.POST.get("pages_verificacion"):
            page_min = (int(request.POST.get("pages_verificacion")) - 1) * 15
            page_max = int(request.POST.get("pages_verificacion")) * 15
            verificaciones = HistorialVerificaciones.objects.select_related("lead").order_by("-id")[page_min:page_max]
            verificaciones = list(verificaciones.values("id", "fecha_hora_verificacion", "lead__prospecto__nombre", "responsable", "lead__prospecto__celular", "lead__origen_lead", "estado_llamada", "lead__nombre_anfitrion", "tipo_solicitud", "lead__sala", "lead__prospecto__nombre_asesor", "lead__comentario"))
            return JsonResponse(verificaciones, safe=False)
        if request.POST.get("pages_agendados"):
            page_min = (int(request.POST.get("pages_agendados")) - 1) * 15
            page_max = int(request.POST.get("pages_agendados")) * 15
            leads_agendados = leads.filter(nombre_asesor__isnull=False, activo=True).exclude(etapa="Desistido").order_by("-id")[page_min:page_max]
            leads_agendados = list(leads_agendados.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion", "tipo_solicitud_verificacion", "prospecto__nombre_asesor"))
            return JsonResponse(leads_agendados, safe=False)
        if request.POST.get("filtrar_primer_contacto"):
            page_min = 0
            page_max = 15
            desde_primer_contacto = request.POST.get("desde_primer_contacto")
            hasta_primer_contacto = request.POST.get("hasta_primer_contacto")

            leads_primer_contacto = Lead.objects.filter(nombre_asesor__isnull=True, activo=True).order_by("-id")
            
            if desde_primer_contacto:
                desde_primer_contacto = datetime.strptime(desde_primer_contacto, '%Y-%m-%d').date()
                leads_primer_contacto = leads_primer_contacto.filter(fecha_apertura__gte=desde_primer_contacto)
            if hasta_primer_contacto:
                hasta_primer_contacto = datetime.strptime(hasta_primer_contacto, '%Y-%m-%d').date()
                leads_primer_contacto = leads_primer_contacto.filter(fecha_apertura__lte=hasta_primer_contacto)

            print(leads_primer_contacto)

            if (leads_primer_contacto.count() % 15) == 0:
                cantidad_filtrado_pag = leads_primer_contacto.count() // 15
            else:
                cantidad_filtrado_pag = leads_primer_contacto.count() // 15 + 1

            leads_primer_contacto = leads_primer_contacto[page_min:page_max]
            
            leads_primer_contacto = list(leads_primer_contacto.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_primer_contacto.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_primer_contacto)
            return JsonResponse(leads_primer_contacto, safe=False)
        if request.POST.get("filtrar_verificacion"):
            page_min = 0
            page_max = 15
            desde_verificacion = request.POST.get("desde_verificacion")
            hasta_verificacion = request.POST.get("hasta_verificacion")

            verificaciones = HistorialVerificaciones.objects.select_related("lead").order_by("-id")
            
            if desde_verificacion:
                desde_verificacion = datetime.strptime(desde_verificacion, '%Y-%m-%d').date()
                verificaciones = verificaciones.filter(fecha_hora_verificacion__gte=desde_verificacion)
            if hasta_verificacion:
                hasta_verificacion = datetime.strptime(hasta_verificacion, '%Y-%m-%d').date()
                verificaciones = verificaciones.filter(fecha_hora_verificacion__lte=hasta_verificacion)

            print(verificaciones)

            if (verificaciones.count() % 15) == 0:
                cantidad_filtrado_pag = verificaciones.count() // 15
            else:
                cantidad_filtrado_pag = verificaciones.count() // 15 + 1

            verificaciones = verificaciones[page_min:page_max]
            
            verificaciones = list(verificaciones.values("id", "fecha_hora_verificacion", "lead__prospecto__nombre", "responsable", "lead__prospecto__celular", "lead__origen_lead", "estado_llamada", "lead__nombre_anfitrion", "tipo_solicitud", "lead__sala", "lead__prospecto__nombre_asesor"))
            verificaciones.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(verificaciones)
            return JsonResponse(verificaciones, safe=False)
        if request.POST.get("filtrar_agendados"):
            page_min = 0
            page_max = 15
            anfitrion_agendados = request.POST.get("anfitrion_agendados")
            verificado_agendados = request.POST.get("verificado_agendados")
            tipo_solicitud_agendados = request.POST.get("tipo_solicitud_agendados")
            asesor_agendados = request.POST.get("asesor_agendados")
            sala_agendados = request.POST.get("sala_agendados")
            desde_agendados = request.POST.get("desde_agendados")
            hasta_agendados = request.POST.get("hasta_agendados")
            search_agendados = request.POST.get("search_agendados")

            leads_agendados = Lead.objects.filter(nombre_asesor__isnull=False, activo=True).exclude(etapa="Desistido").order_by("-id")
            
            if anfitrion_agendados:
                leads_agendados = leads_agendados.filter(nombre_anfitrion=anfitrion_agendados)
            if verificado_agendados:
                if verificado_agendados == "SI":
                    leads_agendados = leads_agendados.filter(estado_llamada_verificacion__isnull=False)
                elif verificado_agendados == "NO":
                    leads_agendados = leads_agendados.filter(estado_llamada_verificacion__isnull=True)
            if tipo_solicitud_agendados:
                leads_agendados = leads_agendados.filter(tipo_solicitud_verificacion=tipo_solicitud_agendados)
            if asesor_agendados:
                leads_agendados = leads_agendados.filter(nombre_asesor=asesor_agendados)
            if sala_agendados:
                leads_agendados = leads_agendados.filter(sala=sala_agendados)
            if desde_agendados:
                desde_agendados = datetime.strptime(desde_agendados, '%Y-%m-%d').date()
                leads_agendados = leads_agendados.filter(fecha_apertura__gte=desde_agendados)
            if hasta_agendados:
                hasta_agendados = datetime.strptime(hasta_agendados, '%Y-%m-%d').date()
                leads_agendados = leads_agendados.filter(fecha_apertura__lte=hasta_agendados)
            if search_agendados:
                leads_agendados = leads_agendados.filter(prospecto__nombre__icontains=search_agendados) | leads_agendados.filter(prospecto__celular__icontains=search_agendados)

            print(leads_agendados)
            
            if (leads_agendados.count() % 15) == 0:
                cantidad_filtrado_pag = leads_agendados.count() // 15
            else:
                cantidad_filtrado_pag = leads_agendados.count() // 15 + 1

            leads_agendados = leads_agendados[page_min:page_max]
            
            leads_agendados = list(leads_agendados.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_agendados.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            return JsonResponse(leads_agendados, safe=False)

class OperativoAsesorView(LoginRequiredMixin, TemplateView):
    # Vista de Operativo Asesor

    template_name = "OperativoAsesor.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)
        
        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads = Lead.objects.filter(activo=True, nombre_asesor__isnull=False).order_by("-id")

        functions.verificar_primer_contacto_todos_los_leads(leads)

        if calendario_general == False:
            leads_no_contactado = Lead.objects.filter(etapa="No contactado", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            leads_interaccion = Lead.objects.filter(etapa="Interaccion", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            leads_oportunidad = Lead.objects.filter(etapa="Oportunidad", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            leads_pedido = Lead.objects.filter(etapa="Pedido", activo=True, nombre_asesor=functions.separar_nombre(user.username)).exclude(respuesta="Entrega finalizada").order_by("-id")
        else:

            leads_no_contactado = Lead.objects.filter(etapa="No contactado", activo=True, nombre_asesor__isnull=False).order_by("-id")
            leads_interaccion = Lead.objects.filter(etapa="Interaccion", activo=True, nombre_asesor__isnull=False).order_by("-id")
            leads_oportunidad = Lead.objects.filter(etapa="Oportunidad", activo=True, nombre_asesor__isnull=False).order_by("-id")
            leads_pedido = Lead.objects.filter(etapa="Pedido", activo=True, nombre_asesor__isnull=False).exclude(respuesta="Entrega finalizada").order_by("-id")
        respuestas = CatalogoRespuestasByEtapa.objects.all()
        estados = CatalogoRespuestasByEtapa.objects.all()

        mostrado_marcas = VehiculosInteresLead.objects.filter(mostrado=True).values("lead", "mostrado").distinct().annotate(latest=Max("id")).values("lead", "marca", "modelo")

        origenes_lead_no_contactado = leads_no_contactado.order_by("origen_lead").values("origen_lead").distinct()
        origenes_lead_interaccion = leads_interaccion.order_by("origen_lead").values("origen_lead").distinct()
        origenes_lead_oportunidad = leads_oportunidad.order_by("origen_lead").values("origen_lead").distinct()
        origenes_lead_pedido = leads_pedido.order_by("origen_lead").values("origen_lead").distinct()

        respuestas_no_contactado = leads_no_contactado.order_by("respuesta").values("respuesta").distinct()
        respuestas_interaccion = leads_interaccion.order_by("respuesta").values("respuesta").distinct()
        respuestas_oportunidad = leads_oportunidad.order_by("respuesta").values("respuesta").distinct()
        respuestas_pedido = leads_pedido.order_by("respuesta").values("respuesta").distinct()

        estados_no_contactado = leads_no_contactado.order_by("estado").values("estado").distinct()
        estados_interaccion = leads_interaccion.order_by("estado").values("estado").distinct()
        estados_oportunidad = leads_oportunidad.order_by("estado").values("estado").distinct()
        estados_pedido = leads_pedido.order_by("estado").values("estado").distinct()

        asesores_no_contactado = leads_no_contactado.order_by("nombre_asesor").values("nombre_asesor").distinct()
        asesores_interaccion = leads_interaccion.order_by("nombre_asesor").values("nombre_asesor").distinct()
        asesores_oportunidad = leads_oportunidad.order_by("nombre_asesor").values("nombre_asesor").distinct()
        asesores_pedido = leads_pedido.order_by("nombre_asesor").values("nombre_asesor").distinct()

        salas_no_contactado = leads_no_contactado.order_by("sala").values("sala").distinct()
        salas_interaccion = leads_interaccion.order_by("sala").values("sala").distinct()
        salas_oportunidad = leads_oportunidad.order_by("sala").values("sala").distinct()
        salas_pedido = leads_pedido.order_by("sala").values("sala").distinct()

        marcas_no_contactado = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_no_contactado).values("marca").distinct()
        marcas_interaccion = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_interaccion).values("marca").distinct()
        marcas_oportunidad = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_oportunidad).values("marca").distinct()
        marcas_pedido = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_pedido).values("marca").distinct()

        modelos_no_contactado = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_no_contactado).values("modelo").distinct() 
        modelos_interaccion = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_interaccion).values("modelo").distinct() 
        modelos_oportunidad = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_oportunidad).values("modelo").distinct() 
        modelos_pedido = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_pedido).values("modelo").distinct() 

        print("marcas_pedido")
        print(marcas_pedido)
        
        if (leads_no_contactado.count() % 15) == 0:
            cantidad_no_contactado_pag = leads_no_contactado.count() // 15
        else:
            cantidad_no_contactado_pag = leads_no_contactado.count() // 15 + 1

        if (leads_interaccion.count() % 15) == 0:
            cantidad_interaccion_pag = leads_interaccion.count() // 15
        else:
            cantidad_interaccion_pag = leads_interaccion.count() // 15 + 1

        if (leads_oportunidad.count() % 15) == 0:
            cantidad_oportunidad_pag = leads_oportunidad.count() // 15
        else:
            cantidad_oportunidad_pag = leads_oportunidad.count() // 15 + 1

        if (leads_pedido.count() % 15) == 0:
            cantidad_pedido_pag = leads_pedido.count() // 15
        else:
            cantidad_pedido_pag = leads_pedido.count() // 15 + 1

        context["asesor_actual"] = asesor_actual
        context["asesores_no_contactado"] = asesores_no_contactado
        context["asesores_interaccion"] = asesores_interaccion
        context["asesores_oportunidad"] = asesores_oportunidad
        context["asesores_pedido"] = asesores_pedido
        context["calendario_general"] = calendario_general
        context["cantidad_no_contactado"] = leads_no_contactado.count()
        context["cantidad_interaccion"] = leads_interaccion.count()
        context["cantidad_oportunidad"] = leads_oportunidad.count()
        context["cantidad_pedido"] = leads_pedido.count()
        context["cantidad_no_contactado_pag"] = cantidad_no_contactado_pag
        context["cantidad_interaccion_pag"] = cantidad_interaccion_pag
        context["cantidad_oportunidad_pag"] = cantidad_oportunidad_pag
        context["cantidad_pedido_pag"] = cantidad_pedido_pag
        context["estados"] = estados
        context["estados_no_contactado"] = estados_no_contactado
        context["estados_interaccion"] = estados_interaccion
        context["estados_oportunidad"] = estados_oportunidad
        context["estados_pedido"] = estados_pedido
        context["leads_no_contactado"] = leads_no_contactado[0:15]
        context["leads_interaccion"] = leads_interaccion[0:15]
        context["leads_oportunidad"] = leads_oportunidad[0:15]
        context["leads_pedido"] = leads_pedido[0:15]
        context["marcas_no_contactado"] = marcas_no_contactado
        context["marcas_interaccion"] = marcas_interaccion
        context["marcas_oportunidad"] = marcas_oportunidad
        context["marcas_pedido"] = marcas_pedido
        context["modelos_no_contactado"] = modelos_no_contactado
        context["modelos_interaccion"] = modelos_interaccion
        context["modelos_oportunidad"] = modelos_oportunidad
        context["modelos_pedido"] = modelos_pedido
        context["mostrado_marcas"] = mostrado_marcas
        context["origenes_lead_no_contactado"] = origenes_lead_no_contactado
        context["origenes_lead_interaccion"] = origenes_lead_interaccion
        context["origenes_lead_oportunidad"] = origenes_lead_oportunidad
        context["origenes_lead_pedido"] = origenes_lead_pedido
        context["pages_no_contactado"] = 1
        context["pages_interaccion"] = 1
        context["pages_oportunidad"] = 1
        context["pages_pedido"] = 1
        context["respuestas"] = respuestas
        context["respuestas_no_contactado"] = respuestas_no_contactado
        context["respuestas_interaccion"] = respuestas_interaccion
        context["respuestas_oportunidad"] = respuestas_oportunidad
        context["respuestas_pedido"] = respuestas_pedido
        context["salas_no_contactado"] = salas_no_contactado
        context["salas_interaccion"] = salas_interaccion
        context["salas_oportunidad"] = salas_oportunidad
        context["salas_pedido"] = salas_pedido
        context["user"] = user

        return context
    
    def post(self, request):

        user = User.objects.get(username=self.request.user)
        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads = Lead.objects.all()

        origen_lead_no_contactado = request.POST.get("origen_lead_no_contactado")
        respuesta_no_contactado = request.POST.get("respuesta_no_contactado")
        estado_no_contactado = request.POST.get("estado_no_contactado")
        asesor_no_contactado = request.POST.get("asesor_no_contactado")
        sala_no_contactado = request.POST.get("sala_no_contactado")
        marca_no_contactado = request.POST.get("marca_no_contactado")
        modelo_no_contactado = request.POST.get("modelo_no_contactado")
        desde_no_contactado = request.POST.get("desde_no_contactado")
        hasta_no_contactado = request.POST.get("hasta_no_contactado")
        search_no_contactado = request.POST.get("search_no_contactado")

        try:
            json.loads(request.POST.get("if_filtrar_no_contactado"))
            if_filtrar_no_contactado = True
        except:
            if_filtrar_no_contactado = False

        if if_filtrar_no_contactado:
            if origen_lead_no_contactado:
                leads = leads.filter(origen_lead=origen_lead_no_contactado)
            if respuesta_no_contactado:
                leads = leads.filter(respuesta=respuesta_no_contactado)
            if estado_no_contactado:
                leads = leads.filter(estado=estado_no_contactado)
            if asesor_no_contactado:
                leads = leads.filter(nombre_asesor=asesor_no_contactado)
            if sala_no_contactado:
                leads = leads.filter(sala=sala_no_contactado)
            if desde_no_contactado:
                desde_no_contactado = datetime.strptime(desde_no_contactado, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_no_contactado)
            if hasta_no_contactado:
                hasta_no_contactado = datetime.strptime(hasta_no_contactado, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_no_contactado)
            if search_no_contactado:
                leads = leads.filter(prospecto__nombre__icontains=search_no_contactado) | leads.filter(prospecto__celular__icontains=search_no_contactado)

        origen_lead_interaccion = request.POST.get("origen_lead_interaccion")
        respuesta_interaccion = request.POST.get("respuesta_interaccion")
        estado_interaccion = request.POST.get("estado_interaccion")
        asesor_interaccion = request.POST.get("asesor_interaccion")
        sala_interaccion = request.POST.get("sala_interaccion")
        marca_interaccion = request.POST.get("marca_interaccion")
        modelo_interaccion = request.POST.get("modelo_interaccion")
        desde_interaccion = request.POST.get("desde_interaccion")
        hasta_interaccion = request.POST.get("hasta_interaccion")
        search_interaccion = request.POST.get("search_interaccion")

        try:
            json.loads(request.POST.get("if_filtrar_interaccion"))
            if_filtrar_interaccion = True
        except:
            if_filtrar_interaccion = False

        if if_filtrar_interaccion:
            if origen_lead_interaccion:
                leads = leads.filter(origen_lead=origen_lead_interaccion)
            if respuesta_interaccion:
                leads = leads.filter(respuesta=respuesta_interaccion)
            if estado_interaccion:
                leads = leads.filter(estado=estado_interaccion)
            if asesor_interaccion:
                leads = leads.filter(nombre_asesor=asesor_interaccion)
            if sala_interaccion:
                leads = leads.filter(sala=sala_interaccion)
            if desde_interaccion:
                desde_interaccion = datetime.strptime(desde_interaccion, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_interaccion)
            if hasta_interaccion:
                hasta_interaccion = datetime.strptime(hasta_interaccion, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_interaccion)
            if search_interaccion:
                leads = leads.filter(prospecto__nombre__icontains=search_interaccion) | leads.filter(prospecto__celular__icontains=search_interaccion)

        origen_lead_oportunidad = request.POST.get("origen_lead_oportunidad")
        respuesta_oportunidad = request.POST.get("respuesta_oportunidad")
        estado_oportunidad = request.POST.get("estado_oportunidad")
        asesor_oportunidad = request.POST.get("asesor_oportunidad")
        sala_oportunidad = request.POST.get("sala_oportunidad")
        marca_oportunidad = request.POST.get("marca_oportunidad")
        modelo_oportunidad = request.POST.get("modelo_oportunidad")
        desde_oportunidad = request.POST.get("desde_oportunidad")
        hasta_oportunidad = request.POST.get("hasta_oportunidad")
        search_oportunidad = request.POST.get("search_oportunidad")

        try:
            json.loads(request.POST.get("if_filtrar_oportunidad"))
            if_filtrar_oportunidad = True
        except:
            if_filtrar_oportunidad = False

        if if_filtrar_oportunidad:
            if origen_lead_oportunidad:
                leads = leads.filter(origen_lead=origen_lead_oportunidad)
            if respuesta_oportunidad:
                leads = leads.filter(respuesta=respuesta_oportunidad)
            if estado_oportunidad:
                leads = leads.filter(estado=estado_oportunidad)
            if asesor_oportunidad:
                leads = leads.filter(nombre_asesor=asesor_oportunidad)
            if sala_oportunidad:
                leads = leads.filter(sala=sala_oportunidad)
            if desde_oportunidad:
                desde_oportunidad = datetime.strptime(desde_oportunidad, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_oportunidad)
            if hasta_oportunidad:
                hasta_oportunidad = datetime.strptime(hasta_oportunidad, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_oportunidad)
            if search_oportunidad:
                leads = leads.filter(prospecto__nombre__icontains=search_oportunidad) | leads.filter(prospecto__celular__icontains=search_oportunidad)

        origen_lead_pedido = request.POST.get("origen_lead_pedido")
        respuesta_pedido = request.POST.get("respuesta_pedido")
        estado_pedido = request.POST.get("estado_pedido")
        asesor_pedido = request.POST.get("asesor_pedido")
        sala_pedido = request.POST.get("sala_pedido")
        marca_pedido = request.POST.get("marca_pedido")
        modelo_pedido = request.POST.get("modelo_pedido")
        desde_pedido = request.POST.get("desde_pedido")
        hasta_pedido = request.POST.get("hasta_pedido")
        search_pedido = request.POST.get("search_pedido")

        try:
            json.loads(request.POST.get("if_filtrar_pedido"))
            if_filtrar_pedido = True
        except:
            if_filtrar_pedido = False

        if if_filtrar_pedido:
            if origen_lead_pedido:
                leads = leads.filter(origen_lead=origen_lead_pedido)
            if respuesta_pedido:
                leads = leads.filter(respuesta=respuesta_pedido)
            if estado_pedido:
                leads = leads.filter(estado=estado_pedido)
            if asesor_pedido:
                leads = leads.filter(nombre_asesor=asesor_pedido)
            if sala_pedido:
                leads = leads.filter(sala=sala_pedido)
            if desde_pedido:
                desde_pedido = datetime.strptime(desde_pedido, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_pedido)
            if hasta_pedido:
                hasta_pedido = datetime.strptime(hasta_pedido, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_pedido)
            if search_pedido:
                leads = leads.filter(prospecto__nombre__icontains=search_pedido) | leads.filter(prospecto__celular__icontains=search_pedido)


        if request.POST.get("pages_no_contactado"):
            page_min = (int(request.POST.get("pages_no_contactado")) - 1) * 15
            page_max = int(request.POST.get("pages_no_contactado")) * 15
            if calendario_general == False:
                leads_no_contactado = leads.filter(etapa="No contactado", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")[page_min:page_max]
            else:
                leads_no_contactado = leads.filter(etapa="No contactado", activo=True, nombre_asesor__isnull=False).order_by("-id")[page_min:page_max]
            leads_no_contactado = list(leads_no_contactado.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_no_contactado, safe=False)
        if request.POST.get("pages_interaccion"):
            page_min = (int(request.POST.get("pages_interaccion")) - 1) * 15
            page_max = int(request.POST.get("pages_interaccion")) * 15
            if calendario_general == False:
                leads_interaccion = leads.filter(etapa="Interaccion", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")[page_min:page_max]
            else:
                leads_interaccion = leads.filter(etapa="Interaccion", activo=True, nombre_asesor__isnull=False).order_by("-id")[page_min:page_max]
            leads_interaccion = list(leads_interaccion.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_interaccion, safe=False)
        if request.POST.get("pages_oportunidad"):
            page_min = (int(request.POST.get("pages_oportunidad")) - 1) * 15
            page_max = int(request.POST.get("pages_oportunidad")) * 15
            if calendario_general == False:
                leads_oportunidad = leads.filter(etapa="Oportunidad", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")[page_min:page_max]
            else:
                leads_oportunidad = leads.filter(etapa="Oportunidad", activo=True, nombre_asesor__isnull=False).order_by("-id")[page_min:page_max]
            leads_oportunidad = list(leads_oportunidad.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_oportunidad, safe=False)
        if request.POST.get("pages_pedido"):
            page_min = (int(request.POST.get("pages_pedido")) - 1) * 15
            page_max = int(request.POST.get("pages_pedido")) * 15
            if calendario_general == False:
                leads_pedido = leads.filter(etapa="Pedido", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")[page_min:page_max]
            else:
                leads_pedido = leads.filter(etapa="Pedido", activo=True, nombre_asesor__isnull=False).order_by("-id")[page_min:page_max]
            leads_pedido = list(leads_pedido.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_pedido, safe=False)
        if request.POST.get("filtrar_no_contactado"):
            page_min = 0
            page_max = 15
            origen_lead_no_contactado = request.POST.get("origen_lead_no_contactado")
            respuesta_no_contactado = request.POST.get("respuesta_no_contactado")
            estado_no_contactado = request.POST.get("estado_no_contactado")
            asesor_no_contactado = request.POST.get("asesor_no_contactado")
            sala_no_contactado = request.POST.get("sala_no_contactado")
            marca_no_contactado = request.POST.get("marca_no_contactado")
            modelo_no_contactado = request.POST.get("modelo_no_contactado")
            desde_no_contactado = request.POST.get("desde_no_contactado")
            hasta_no_contactado = request.POST.get("hasta_no_contactado")
            search_no_contactado = request.POST.get("search_no_contactado")

            if calendario_general == False:
                leads_no_contactado = Lead.objects.filter(etapa="No contactado", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            else:
                leads_no_contactado = Lead.objects.filter(etapa="No contactado", activo=True, nombre_asesor__isnull=False).order_by("-id")
            
            if origen_lead_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(origen_lead=origen_lead_no_contactado)
            if respuesta_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(respuesta=respuesta_no_contactado)
            if estado_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(estado=estado_no_contactado)
            if asesor_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(nombre_asesor=asesor_no_contactado)
            if sala_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(sala=sala_no_contactado)
            if desde_no_contactado:
                desde_no_contactado = datetime.strptime(desde_no_contactado, '%Y-%m-%d').date()
                leads_no_contactado = leads_no_contactado.filter(fecha_apertura__gte=desde_no_contactado)
            if hasta_no_contactado:
                hasta_no_contactado = datetime.strptime(hasta_no_contactado, '%Y-%m-%d').date()
                leads_no_contactado = leads_no_contactado.filter(fecha_apertura__lte=hasta_no_contactado)
            if search_no_contactado:
                leads_no_contactado = leads_no_contactado.filter(prospecto__nombre__icontains=search_no_contactado) | leads_no_contactado.filter(prospecto__celular__icontains=search_no_contactado)

            if (leads_no_contactado.count() % 15) == 0:
                cantidad_filtrado_pag = leads_no_contactado.count() // 15
            else:
                cantidad_filtrado_pag = leads_no_contactado.count() // 15 + 1


            leads_no_contactado = leads_no_contactado[page_min:page_max]
            
            leads_no_contactado = list(leads_no_contactado.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_no_contactado.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_no_contactado)
            return JsonResponse(leads_no_contactado, safe=False)
        
        if request.POST.get("filtrar_interaccion"):
            page_min = 0
            page_max = 15
            origen_lead_interaccion = request.POST.get("origen_lead_interaccion")
            respuesta_interaccion = request.POST.get("respuesta_interaccion")
            estado_interaccion = request.POST.get("estado_interaccion")
            asesor_interaccion = request.POST.get("asesor_interaccion")
            sala_interaccion = request.POST.get("sala_interaccion")
            marca_interaccion = request.POST.get("marca_interaccion")
            modelo_interaccion = request.POST.get("modelo_interaccion")
            desde_interaccion = request.POST.get("desde_interaccion")
            hasta_interaccion = request.POST.get("hasta_interaccion")
            search_interaccion = request.POST.get("search_interaccion")

            if calendario_general == False:
                leads_interaccion = Lead.objects.filter(etapa="Interaccion", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            else:
                leads_interaccion = Lead.objects.filter(etapa="Interaccion", activo=True, nombre_asesor__isnull=False).order_by("-id")
            
            if origen_lead_interaccion:
                leads_interaccion = leads_interaccion.filter(origen_lead=origen_lead_interaccion)
            if respuesta_interaccion:
                leads_interaccion = leads_interaccion.filter(respuesta=respuesta_interaccion)
            if estado_interaccion:
                leads_interaccion = leads_interaccion.filter(estado=estado_interaccion)
            if asesor_interaccion:
                leads_interaccion = leads_interaccion.filter(nombre_asesor=asesor_interaccion)
            if sala_interaccion:
                leads_interaccion = leads_interaccion.filter(sala=sala_interaccion)
            if desde_interaccion:
                desde_interaccion = datetime.strptime(desde_interaccion, '%Y-%m-%d').date()
                leads_interaccion = leads_interaccion.filter(fecha_apertura__gte=desde_interaccion)
            if hasta_interaccion:
                hasta_interaccion = datetime.strptime(hasta_interaccion, '%Y-%m-%d').date()
                leads_interaccion = leads_interaccion.filter(fecha_apertura__lte=hasta_interaccion)
            if search_interaccion:
                leads_interaccion = leads_interaccion.filter(prospecto__nombre__icontains=search_interaccion) | leads_interaccion.filter(prospecto__celular__icontains=search_interaccion)

            if (leads_interaccion.count() % 15) == 0:
                cantidad_filtrado_pag = leads_interaccion.count() // 15
            else:
                cantidad_filtrado_pag = leads_interaccion.count() // 15 + 1

            leads_interaccion = leads_interaccion[page_min:page_max]
            
            leads_interaccion = list(leads_interaccion.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_interaccion.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_interaccion)
            return JsonResponse(leads_interaccion, safe=False)
        
        if request.POST.get("filtrar_oportunidad"):
            page_min = 0
            page_max = 15
            origen_lead_oportunidad = request.POST.get("origen_lead_oportunidad")
            respuesta_oportunidad = request.POST.get("respuesta_oportunidad")
            estado_oportunidad = request.POST.get("estado_oportunidad")
            asesor_oportunidad = request.POST.get("asesor_oportunidad")
            sala_oportunidad = request.POST.get("sala_oportunidad")
            marca_oportunidad = request.POST.get("marca_oportunidad")
            modelo_oportunidad = request.POST.get("modelo_oportunidad")
            desde_oportunidad = request.POST.get("desde_oportunidad")
            hasta_oportunidad = request.POST.get("hasta_oportunidad")
            search_oportunidad = request.POST.get("search_oportunidad")

            if calendario_general == False:
                leads_oportunidad = Lead.objects.filter(etapa="Oportunidad", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            else:
                leads_oportunidad = Lead.objects.filter(etapa="Oportunidad", activo=True, nombre_asesor__isnull=False).order_by("-id")
            
            if origen_lead_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(origen_lead=origen_lead_oportunidad)
            if respuesta_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(respuesta=respuesta_oportunidad)
            if estado_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(estado=estado_oportunidad)
            if asesor_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(nombre_asesor=asesor_oportunidad)
            if sala_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(sala=sala_oportunidad)
            if desde_oportunidad:
                desde_oportunidad = datetime.strptime(desde_oportunidad, '%Y-%m-%d').date()
                leads_oportunidad = leads_oportunidad.filter(fecha_apertura__gte=desde_oportunidad)
            if hasta_oportunidad:
                hasta_oportunidad = datetime.strptime(hasta_oportunidad, '%Y-%m-%d').date()
                leads_oportunidad = leads_oportunidad.filter(fecha_apertura__lte=hasta_oportunidad)
            if search_oportunidad:
                leads_oportunidad = leads_oportunidad.filter(prospecto__nombre__icontains=search_oportunidad) | leads_oportunidad.filter(prospecto__celular__icontains=search_oportunidad)

            if (leads_oportunidad.count() % 15) == 0:
                cantidad_filtrado_pag = leads_oportunidad.count() // 15
            else:
                cantidad_filtrado_pag = leads_oportunidad.count() // 15 + 1

            leads_oportunidad = leads_oportunidad[page_min:page_max]
            
            leads_oportunidad = list(leads_oportunidad.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_oportunidad.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_oportunidad)
            return JsonResponse(leads_oportunidad, safe=False)
        
        if request.POST.get("filtrar_pedido"):
            page_min = 0
            page_max = 15
            origen_lead_pedido = request.POST.get("origen_lead_pedido")
            respuesta_pedido = request.POST.get("respuesta_pedido")
            estado_pedido = request.POST.get("estado_pedido")
            asesor_pedido = request.POST.get("asesor_pedido")
            sala_pedido = request.POST.get("sala_pedido")
            marca_pedido = request.POST.get("marca_pedido")
            modelo_pedido = request.POST.get("modelo_pedido")
            desde_pedido = request.POST.get("desde_pedido")
            hasta_pedido = request.POST.get("hasta_pedido")
            search_pedido = request.POST.get("search_pedido")

            if calendario_general == False:
                leads_pedido = Lead.objects.filter(etapa="Pedido", activo=True, nombre_asesor=functions.separar_nombre(user.username)).order_by("-id")
            else:
                leads_pedido = Lead.objects.filter(etapa="Pedido", activo=True, nombre_asesor__isnull=False).order_by("-id")
            
            if origen_lead_pedido:
                leads_pedido = leads_pedido.filter(origen_lead=origen_lead_pedido)
            if respuesta_pedido:
                leads_pedido = leads_pedido.filter(respuesta=respuesta_pedido)
            if estado_pedido:
                leads_pedido = leads_pedido.filter(estado=estado_pedido)
            if asesor_pedido:
                leads_pedido = leads_pedido.filter(nombre_asesor=asesor_pedido)
            if sala_pedido:
                leads_pedido = leads_pedido.filter(sala=sala_pedido)
            if desde_pedido:
                desde_pedido = datetime.strptime(desde_pedido, '%Y-%m-%d').date()
                leads_pedido = leads_pedido.filter(fecha_apertura__gte=desde_pedido)
            if hasta_pedido:
                hasta_pedido = datetime.strptime(hasta_pedido, '%Y-%m-%d').date()
                leads_pedido = leads_pedido.filter(fecha_apertura__lte=hasta_pedido)
            if search_pedido:
                leads_pedido = leads_pedido.filter(prospecto__nombre__icontains=search_pedido) | leads_pedido.filter(prospecto__celular__icontains=search_pedido)

            if (leads_pedido.count() % 15) == 0:
                cantidad_filtrado_pag = leads_pedido.count() // 15
            else:
                cantidad_filtrado_pag = leads_pedido.count() // 15 + 1

            leads_pedido = leads_pedido[page_min:page_max]
            
            leads_pedido = list(leads_pedido.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__apellido_paterno",  "prospecto__apellido_materno", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_pedido.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_pedido)
            return JsonResponse(leads_pedido, safe=False)
    
    
class ReportesView(LoginRequiredMixin, TemplateView):
    # Vista de Reportes

    template_name = "Reports.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)

        leads = Lead.objects.filter(activo=True, nombre_asesor__isnull=False).order_by("-id")

        functions.verificar_primer_contacto_todos_los_leads(leads)

        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads_agendados = Lead.objects.filter(nombre_asesor__isnull=False, activo=True).order_by("-id")
        separados_y_facturados = VehiculosInteresLead.objects.filter(Q(separado=True) | Q(facturado=True)).values_list("lead").distinct()
        separados = VehiculosInteresLead.objects.filter(separado=True).values_list("lead").distinct()
        facturados = VehiculosInteresLead.objects.filter(facturado=True).values_list("lead").distinct()
        leads_facturados = Lead.objects.filter(pk__in=facturados, activo=True).order_by("-id")
        leads_separados = Lead.objects.filter(pk__in=separados, activo=True).order_by("-id")
        leads_separados_y_facturados = Lead.objects.filter(pk__in=separados_y_facturados, activo=True).order_by("-id")
        leads_desistidos = Lead.objects.filter(etapa="Desistido", activo=True).order_by("-id")
        leads_concretados = Lead.objects.filter(etapa="Pedido", respuesta="Entrega finalizada", activo=True).order_by("-id")

        historial = Historial.objects.filter(lead__in=leads_agendados).values("lead").annotate(Max("fecha"))

        print("historial")
        print(historial)

        verificados = HistorialVerificaciones.objects.values("lead", "tipo_solicitud").distinct().order_by("-id")

        mostrado_marcas = VehiculosInteresLead.objects.filter(mostrado=True).values("lead").distinct().values("lead", "marca", "modelo", "codigo_vehiculo")
        separados_y_facturados_marcas = VehiculosInteresLead.objects.filter(mostrado=False).values("lead").distinct().values("lead", "marca", "modelo", "codigo_vehiculo", "fecha")

        origenes_lead_agendados = leads_agendados.order_by("origen_lead").values("origen_lead").distinct()
        
        respuestas_agendados = leads_agendados.order_by("respuesta").values("respuesta").distinct()
        respuestas_desistidos = leads_desistidos.order_by("respuesta").values("respuesta").distinct()
        
        estados_agendados = leads_agendados.order_by("estado").values("estado").distinct()
        estados_separados_y_facturados = leads_separados_y_facturados.order_by("estado").values("estado").distinct()
        estados_desistidos = leads_desistidos.order_by("estado").values("estado").distinct()

        anfitriones_agendados = leads_agendados.order_by("nombre_anfitrion").values("nombre_anfitrion").distinct()

        asesores_agendados = leads_agendados.order_by("nombre_asesor").values("nombre_asesor").distinct()
        asesores_separados_y_facturados = leads_separados_y_facturados.order_by("nombre_asesor").values("nombre_asesor").distinct()
        asesores_desistidos = leads_desistidos.order_by("nombre_asesor").values("nombre_asesor").distinct()

        salas_agendados = leads_agendados.order_by("sala").values("sala").distinct()
        salas_separados_y_facturados = leads_separados_y_facturados.order_by("sala").values("sala").distinct()
        salas_desistidos = leads_desistidos.order_by("sala").values("sala").distinct()

        mostrado_marcas_agendados = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_agendados).values("lead", "mostrado").distinct().annotate(latest=Max("id")).values("lead", "marca", "modelo")
        mostrado_marcas_separados_y_facturados = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_separados_y_facturados).values("lead", "mostrado").distinct().annotate(latest=Max("id")).values("lead", "marca", "modelo")
        mostrado_marcas_desistidos = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_desistidos).values("lead", "mostrado").distinct().annotate(latest=Max("id")).values("lead", "marca", "modelo")
       
        if (leads_agendados.count() % 15) == 0:
            cantidad_capturados_pag = leads_agendados.count() // 15
        else:
            cantidad_capturados_pag = leads_agendados.count() // 15 + 1
        if (leads_separados_y_facturados.count() % 15) == 0:
            cantidad_sep_y_fac_pag = leads_separados_y_facturados.count() // 15
        else:
            cantidad_sep_y_fac_pag = leads_separados_y_facturados.count() // 15 + 1
        if (leads_desistidos.count() % 15) == 0:
            cantidad_desistidos_pag = leads_desistidos.count() // 15
        else:
            cantidad_desistidos_pag = leads_desistidos.count() // 15 + 1

        context["anfitriones_agendados"] = anfitriones_agendados
        context["asesor_actual"] = asesor_actual
        context["asesores_agendados"] = asesores_agendados
        context["asesores_separados_y_facturados"] = asesores_separados_y_facturados
        context["asesores_desistidos"] = asesores_desistidos
        context["calendario_general"] = calendario_general
        context["cantidad_agendados"] = leads_agendados.count()
        context["cantidad_capturados_pag"] = cantidad_capturados_pag
        context["cantidad_sep_y_fac_pag"] = cantidad_sep_y_fac_pag
        context["cantidad_desistidos_pag"] = cantidad_desistidos_pag
        context["cantidad_concretados"] = leads_concretados.count()
        context["cantidad_desistidos"] = leads_desistidos.count()
        context["cantidad_historial"] = historial.count()
        context["cantidad_separados_y_facturados"] = leads_separados_y_facturados.count()
        context["estados_agendados"] = estados_agendados
        context["estados_separados_y_facturados"] = estados_separados_y_facturados
        context["estados_desistidos"] = estados_desistidos
        context["historial"] = historial
        context["leads_agendados"] = leads_agendados[0:15]
        context["leads_concretados"] = leads_concretados[0:15]
        context["leads_desistidos"] = leads_desistidos[0:15]
        context["leads_facturados"] = leads_facturados[0:15]
        context["leads_separados"] = leads_separados[0:15]
        context["leads_separados_y_facturados"] = leads_separados_y_facturados[0:15]
        context["mostrado_marcas"] = mostrado_marcas
        context["mostrado_marcas_agendados"] = mostrado_marcas_agendados
        context["mostrado_marcas_separados_y_facturados"] = mostrado_marcas_separados_y_facturados
        context["mostrado_marcas_desistidos"] = mostrado_marcas_desistidos
        context["pages_capturados"] = 1
        context["pages_sep_y_fac"] = 1
        context["pages_desistidos"] = 1
        context["origenes_lead_agendados"] = origenes_lead_agendados
        context["respuestas_agendados"] = respuestas_agendados
        context["respuestas_desistidos"] = respuestas_desistidos
        context["salas_agendados"] = salas_agendados
        context["salas_separados_y_facturados"] = salas_separados_y_facturados
        context["salas_desistidos"] = salas_desistidos
        context["separados"] = separados
        context["separados_y_facturados_marcas"] = separados_y_facturados_marcas
        context["user"] = user
        context["verificados"] = verificados

        return context
    
    def post(self, request):

        leads = Lead.objects.filter(activo=True, nombre_asesor__isnull=False).order_by("-id")

        origen_lead_capturados = request.POST.get("origen_lead_capturados")
        respuesta_capturados = request.POST.get("respuesta_capturados")
        estado_capturados = request.POST.get("estado_capturados")
        anfitrion_capturados = request.POST.get("anfitrion_capturados")
        asesor_capturados = request.POST.get("asesor_capturados")
        sala_capturados = request.POST.get("sala_capturados")
        verificado_capturados = request.POST.get("verificado_capturados")
        marca_capturados = request.POST.get("marca_capturados")
        modelo_capturados = request.POST.get("modelo_capturados")
        desde_capturados = request.POST.get("desde_capturados")
        hasta_capturados = request.POST.get("hasta_capturados")
        search_capturados = request.POST.get("search_capturados")

        try:
            json.loads(request.POST.get("if_filtrar_capturados"))
            if_filtrar_capturados = True
        except:
            if_filtrar_capturados = False


        if if_filtrar_capturados:
            if origen_lead_capturados:
                leads = leads.filter(origen_lead=origen_lead_capturados)
            if respuesta_capturados:
                leads = leads.filter(respuesta=respuesta_capturados)
            if estado_capturados:
                leads = leads.filter(estado=estado_capturados)
            if anfitrion_capturados:
                leads = leads.filter(nombre_asesor=anfitrion_capturados)
            if asesor_capturados:
                leads = leads.filter(nombre_asesor=asesor_capturados)
            if sala_capturados:
                leads = leads.filter(sala=sala_capturados)
            if verificado_capturados:
                if verificado_capturados == "SI":
                    leads = leads.filter(estado_llamada_verificacion__isnull=False)
                elif verificado_capturados == "NO":
                    leads = leads.filter(estado_llamada_verificacion__isnull=True)
            if desde_capturados:
                desde_capturados = datetime.strptime(desde_capturados, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_capturados)
            if hasta_capturados:
                hasta_capturados = datetime.strptime(hasta_capturados, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_capturados)
            if search_capturados:
                leads = leads.filter(prospecto__nombre__icontains=search_capturados) | leads.filter(prospecto__celular__icontains=search_capturados)

        respuesta_desistidos = request.POST.get("respuesta_desistidos")
        estado_desistidos = request.POST.get("estado_desistidos")
        asesor_desistidos = request.POST.get("asesor_desistidos")
        sala_desistidos = request.POST.get("sala_desistidos")
        marca_desistidos = request.POST.get("marca_desistidos")
        modelo_desistidos = request.POST.get("modelo_desistidos")
        desde_desistidos = request.POST.get("desde_desistidos")
        hasta_desistidos = request.POST.get("hasta_desistidos")
        search_desistidos = request.POST.get("search_desistidos")

        try:
            json.loads(request.POST.get("if_filtrar_desistidos"))
            if_filtrar_desistidos = True
        except:
            if_filtrar_desistidos = False


        if if_filtrar_desistidos:
            if respuesta_desistidos:
                leads = leads.filter(respuesta=respuesta_desistidos)
            if estado_desistidos:
                leads = leads.filter(estado=estado_desistidos)
            if asesor_desistidos:
                leads = leads.filter(nombre_asesor=asesor_desistidos)
            if sala_desistidos:
                leads = leads.filter(sala=sala_desistidos)
            if desde_desistidos:
                desde_desistidos = datetime.strptime(desde_desistidos, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__gte=desde_desistidos)
            if hasta_desistidos:
                hasta_desistidos = datetime.strptime(hasta_desistidos, '%Y-%m-%d').date()
                leads = leads.filter(fecha_apertura__lte=hasta_desistidos)
            if search_desistidos:
                leads = leads.filter(prospecto__nombre__icontains=search_desistidos) | leads.filter(prospecto__celular__icontains=search_desistidos)

        if request.POST.get("pages_capturados"):
            page_min = (int(request.POST.get("pages_capturados")) - 1) * 15
            page_max = int(request.POST.get("pages_capturados")) * 15
            leads_agendados = leads.filter(nombre_asesor__isnull=False, activo=True).order_by("-id")[page_min:page_max]
            leads_agendados = list(leads_agendados.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            print(leads_agendados)
            return JsonResponse(leads_agendados, safe=False)
        if request.POST.get("pages_sep_y_fac"):
            page_min = (int(request.POST.get("pages_sep_y_fac")) - 1) * 15
            page_max = int(request.POST.get("pages_sep_y_fac")) * 15
            separados_y_facturados = VehiculosInteresLead.objects.filter(Q(separado=True) | Q(facturado=True)).values_list("lead").distinct()
            leads_separados_y_facturados = leads.filter(pk__in=separados_y_facturados, activo=True).order_by("-id")[page_min:page_max]
            leads_separados_y_facturados = list(leads_separados_y_facturados.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_separados_y_facturados, safe=False)
        if request.POST.get("pages_desistidos"):
            page_min = (int(request.POST.get("pages_desistidos")) - 1) * 15
            page_max = int(request.POST.get("pages_desistidos")) * 15
            leads_desistidos = leads.filter(etapa="Desistido", activo=True).order_by("-id")[page_min:page_max]
            leads_desistidos = list(leads_desistidos.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            print(page_min)
            print(page_max)
            return JsonResponse(leads_desistidos, safe=False)
        
        if request.POST.get("filtrar_capturados"):
            page_min = 0
            page_max = 15
            origen_lead_capturados = request.POST.get("origen_lead_capturados")
            respuesta_capturados = request.POST.get("respuesta_capturados")
            estado_capturados = request.POST.get("estado_capturados")
            anfitrion_capturados = request.POST.get("anfitrion_capturados")
            asesor_capturados = request.POST.get("asesor_capturados")
            sala_capturados = request.POST.get("sala_capturados")
            verificado_capturados = request.POST.get("verificado_capturados")
            marca_capturados = request.POST.get("marca_capturados")
            modelo_capturados = request.POST.get("modelo_capturados")
            desde_capturados = request.POST.get("desde_capturados")
            hasta_capturados = request.POST.get("hasta_capturados")
            search_capturados = request.POST.get("search_capturados")

            leads_capturados = Lead.objects.filter(nombre_asesor__isnull=False, activo=True).order_by("-id")
            
            if origen_lead_capturados:
                leads_capturados = leads_capturados.filter(origen_lead=origen_lead_capturados)
            if respuesta_capturados:
                leads_capturados = leads_capturados.filter(respuesta=respuesta_capturados)
            if estado_capturados:
                leads_capturados = leads_capturados.filter(estado=estado_capturados)
            if anfitrion_capturados:
                leads_capturados = leads_capturados.filter(nombre_anfitrion=anfitrion_capturados)
            if asesor_capturados:
                leads_capturados = leads_capturados.filter(nombre_asesor=asesor_capturados)
            if sala_capturados:
                leads_capturados = leads_capturados.filter(sala=sala_capturados)
            if verificado_capturados:
                if verificado_capturados == "SI":
                    leads_capturados = leads_capturados.filter(estado_llamada_verificacion__isnull=False)
                else:
                    leads_capturados = leads_capturados.filter(estado_llamada_verificacion__isnull=True)
            if desde_capturados:
                desde_capturados = datetime.strptime(desde_capturados, '%Y-%m-%d').date()
                leads_capturados = leads_capturados.filter(fecha_apertura__gte=desde_capturados)
            if hasta_capturados:
                hasta_capturados = datetime.strptime(hasta_capturados, '%Y-%m-%d').date()
                leads_capturados = leads_capturados.filter(fecha_apertura__lte=hasta_capturados)
            if search_capturados:
                leads_capturados = leads_capturados.filter(prospecto__nombre__icontains=search_capturados) | leads_capturados.filter(prospecto__celular__icontains=search_capturados)

            if (leads_capturados.count() % 15) == 0:
                cantidad_filtrado_pag = leads_capturados.count() // 15
            else:
                cantidad_filtrado_pag = leads_capturados.count() // 15 + 1

            leads_capturados = leads_capturados[page_min:page_max]
            
            leads_capturados = list(leads_capturados.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_capturados.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_capturados)
            return JsonResponse(leads_capturados, safe=False)
        
        if request.POST.get("filtrar_desistidos"):
            page_min = 0
            page_max = 15
            origen_lead_desistidos = request.POST.get("origen_lead_desistidos")
            respuesta_desistidos = request.POST.get("respuesta_desistidos")
            estado_desistidos = request.POST.get("estado_desistidos")
            anfitrion_desistidos = request.POST.get("anfitrion_desistidos")
            asesor_desistidos = request.POST.get("asesor_desistidos")
            sala_desistidos = request.POST.get("sala_desistidos")
            verificado_desistidos = request.POST.get("verificado_desistidos")
            marca_desistidos = request.POST.get("marca_desistidos")
            modelo_desistidos = request.POST.get("modelo_desistidos")
            desde_desistidos = request.POST.get("desde_desistidos")
            hasta_desistidos = request.POST.get("hasta_desistidos")
            search_desistidos = request.POST.get("search_desistidos")

            leads_desistidos = Lead.objects.filter(etapa="Desistido", activo=True).order_by("-id")
            
            if origen_lead_desistidos:
                leads_desistidos = leads_desistidos.filter(origen_lead=origen_lead_desistidos)
            if respuesta_desistidos:
                leads_desistidos = leads_desistidos.filter(respuesta=respuesta_desistidos)
            if estado_desistidos:
                leads_desistidos = leads_desistidos.filter(estado=estado_desistidos)
            if anfitrion_desistidos:
                leads_desistidos = leads_desistidos.filter(nombre_anfitrion=anfitrion_desistidos)
            if asesor_desistidos:
                leads_desistidos = leads_desistidos.filter(nombre_asesor=asesor_desistidos)
            if sala_desistidos:
                leads_desistidos = leads_desistidos.filter(sala=sala_desistidos)
            if verificado_desistidos:
                if verificado_desistidos == "SI":
                    leads_desistidos = leads_desistidos.filter(estado_llamada_verificacion__isnull=False)
                else:
                    leads_desistidos = leads_desistidos.filter(estado_llamada_verificacion__isnull=True)
            if desde_desistidos:
                desde_desistidos = datetime.strptime(desde_desistidos, '%Y-%m-%d').date()
                leads_desistidos = leads_desistidos.filter(fecha_apertura__gte=desde_desistidos)
            if hasta_desistidos:
                hasta_desistidos = datetime.strptime(hasta_desistidos, '%Y-%m-%d').date()
                leads_desistidos = leads_desistidos.filter(fecha_apertura__lte=hasta_desistidos)
            if search_desistidos:
                leads_desistidos = leads_desistidos.filter(prospecto__nombre__icontains=search_desistidos) | leads_desistidos.filter(prospecto__celular__icontains=search_desistidos)


            if (leads_desistidos.count() % 15) == 0:
                cantidad_filtrado_pag = leads_desistidos.count() // 15
            else:
                cantidad_filtrado_pag = leads_desistidos.count() // 15 + 1

            leads_desistidos = leads_desistidos[page_min:page_max]
            
            leads_desistidos = list(leads_desistidos.values("id", "fecha_apertura", "prospecto__nombre", "prospecto__celular", "prospecto__correo", "nombre_anfitrion", "tipo_documento", "documento", "campania", "respuesta", "estado", "origen_lead", "sala", "nombre_asesor", "estado_llamada_verificacion"))
            leads_desistidos.append(cantidad_filtrado_pag)
            print(page_min)
            print(page_max)
            print(leads_desistidos)
            return JsonResponse(leads_desistidos, safe=False)
    
class TiemposView(LoginRequiredMixin, TemplateView):
    # Vista de Tiempos

    template_name = "Tiempos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads_tiempos = Lead.objects.filter(tiempo_primer_contacto__isnull=False, activo=True).annotate(dias_totales=(ExpressionWrapper((Cast(datetime.now(), output_field=DateTimeField())) - F('fecha_hora_asignacion_asesor'), output_field=IntegerField())) / 86400000000).annotate(minutos_totales=F("dias_totales") * 1440).order_by("-id")
        leads_verificados = Lead.objects.filter(estado_llamada_verificacion__isnull=False, activo=True).order_by("-id")

        verificados = HistorialVerificaciones.objects.values("lead", "tipo_solicitud").distinct().order_by("-id")

        mostrado_marcas = VehiculosInteresLead.objects.filter(mostrado=True).values("lead").distinct().values("lead", "marca", "modelo")

        origenes_lead_tiempos = leads_tiempos.order_by("origen_lead").values("origen_lead").distinct()
        respuestas_tiempos = leads_tiempos.order_by("respuesta").values("respuesta").distinct()
        estados_tiempos = leads_tiempos.order_by("estado").values("estado").distinct()
        asesores_tiempos = leads_tiempos.order_by("nombre_asesor").values("nombre_asesor").distinct()
        salas_tiempos = leads_tiempos.order_by("sala").values("sala").distinct()
        marcas_tiempos = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_tiempos).values("marca").distinct()
        modelos_tiempos = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_tiempos).values("modelo").distinct() 

        context["asesor_actual"] = asesor_actual
        context["asesores_tiempos"] = asesores_tiempos
        context["calendario_general"] = calendario_general
        context["cantidad_tiempos"] = leads_tiempos.count()
        context["cantidad_verificados"] = leads_verificados.count()
        context["estados_tiempos"] = estados_tiempos
        context["leads_tiempos"] = leads_tiempos
        context["leads_verificados"] = leads_verificados
        context["marcas_tiempos"] = marcas_tiempos
        context["modelos_tiempos"] = modelos_tiempos
        context["mostrado_marcas"] = mostrado_marcas
        context["origenes_lead_tiempos"] = origenes_lead_tiempos
        context["respuestas_tiempos"] = respuestas_tiempos
        context["salas_tiempos"] = salas_tiempos
        context["user"] = user
        context["verificados"] = verificados

        return context
        
class AnuladosView(LoginRequiredMixin, TemplateView):
    # Vista de Anulados

    template_name = "Anulados.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads_anulados = Lead.objects.filter(activo=False).order_by("-id")
        leads_verificados = Lead.objects.filter(estado_llamada_verificacion__isnull=False).order_by("-id")

        verificados = HistorialVerificaciones.objects.values("lead", "tipo_solicitud").distinct().order_by("-id")

        mostrado_marcas = VehiculosInteresLead.objects.filter(mostrado=True).values("lead").distinct().values("lead", "marca", "modelo")


        origenes_lead_anulados = leads_anulados.order_by("origen_lead").values("origen_lead").distinct()
        respuestas_anulados = leads_anulados.order_by("respuesta").values("respuesta").distinct()
        estados_anulados = leads_anulados.order_by("estado").values("estado").distinct()
        asesores_anulados = leads_anulados.order_by("nombre_asesor").values("nombre_asesor").distinct()
        salas_anulados = leads_anulados.order_by("sala").values("sala").distinct()
        marcas_anulados = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_anulados).values("marca").distinct()
        modelos_anulados = VehiculosInteresLead.objects.filter(mostrado=True).filter(lead__in=leads_anulados).values("modelo").distinct() 
        
        context["asesor_actual"] = asesor_actual
        context["asesores_anulados"] = asesores_anulados
        context["calendario_general"] = calendario_general
        context["cantidad_anulados"] = leads_anulados.count()
        context["cantidad_verificados"] = leads_verificados.count()
        context["estados_anulados"] = estados_anulados
        context["leads_anulados"] = leads_anulados
        context["leads_verificados"] = leads_verificados
        context["marcas_anulados"] = marcas_anulados
        context["modelos_anulados"] = modelos_anulados
        context["mostrado_marcas"] = mostrado_marcas
        context["origenes_lead_anulados"] = origenes_lead_anulados
        context["respuestas_anulados"] = respuestas_anulados
        context["salas_anulados"] = salas_anulados
        context["user"] = user
        context["verificados"] = verificados

        return context
    
class ModernizeView(LoginRequiredMixin, TemplateView):
    # Vista de Modernize

    template_name = "Modernize.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        print(user)

        context["user"] = user

        return context
    
class Modernize2View(LoginRequiredMixin, TemplateView):
    # Vista de Modernize2

    template_name = "Modernize2.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        print(user)

        context["user"] = user

        return context
    
class ReportesEventosView(LoginRequiredMixin, TemplateView):
    # Vista de Reportes Eventos

    template_name = "ReportesEventos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)

        leads = Lead.objects.filter(activo=True, nombre_asesor__isnull=False).order_by("-id")

        functions.verificar_primer_contacto_todos_los_leads(leads)

        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0
        print(user)
        
        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        leads_activos = Lead.objects.filter(activo=True, nombre_asesor__isnull=False)
        asesores = Asesor.objects.all()
        asesores_127 = Asesor.objects.filter(sala="127")
        asesores_morato = Asesor.objects.filter(sala="Morato")
        today_min = datetime.combine(timezone.now().date(), datetime.today().time().min)
        today_max = datetime.combine(timezone.now().date(), datetime.today().time().max)
        cantidad_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).count()
        eventos_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).values("asesor").annotate(cantidad=Count("pk"))
        cantidad_pendiente = Evento.objects.filter(fecha_hora__gt=today_max).count()
        eventos_pendientes = Evento.objects.filter(fecha_hora__gt=today_max).values("asesor").annotate(cantidad=Count("pk"))
        cantidad_cumplidos = Evento.objects.filter(fecha_hora__gt=today_max).count()
        eventos_cumplidos = Evento.objects.filter(cumplido=True).values("asesor").annotate(cantidad=Count("pk"))

        print(asesores)
        print("eventos_hoy")
        print(eventos_hoy)
        print(eventos_hoy.values("asesor"))

        print("eventos_cumplidos")
        print(eventos_cumplidos)

        context["asesor_actual"] = asesor_actual
        context["asesores"] = asesores
        context["asesores_127"] = asesores_127
        context["asesores_morato"] = asesores_morato
        context["calendario_general"] = calendario_general
        context["cantidad_activos"] = leads_activos.count()
        context["cantidad_cumplidos"] = cantidad_cumplidos
        context["cantidad_hoy"] = cantidad_hoy
        context["cantidad_pendientes"] = cantidad_pendiente
        context["eventos_cumplidos"] = eventos_cumplidos
        context["eventos_hoy"] = eventos_hoy
        context["eventos_pendientes"] = eventos_pendientes
        context["leads_activos"] = leads_activos
        context["user"] = user

        return context
    
    def post(self, request):

        user = User.objects.get(username=self.request.user)
        if request.POST.get("EventoNombre"):

            nombre = request.POST.get("EventoNombre")
            tipo = request.POST.get("EventoTipo")
            telefono_cliente = request.POST.get("EventoTelefono")
            observaciones = request.POST.get("EventoObservaciones")
            asesor = request.POST.get("EventoAsesor")
            fecha_hora = request.POST.get("EventoFechaHora")
            
            evento = Evento.objects.create(nombre=nombre,
                                           tipo=tipo,
                                           telefono_cliente=telefono_cliente,
                                           observaciones=observaciones,
                                           asesor=Asesor.objects.get(nombre=asesor),
                                           fecha_hora=datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M"),
                                           lead=Lead.objects.get(id=21)
                                           )
            return JsonResponse(evento.pk, safe=False)

class CalendarView(LoginRequiredMixin, TemplateView):
    # Vista de Calendar

    template_name = "Calendar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        asesores = Asesor.objects.all()

        eventos = Evento.objects.all()

        today_min = datetime.combine(timezone.now().date(), datetime.today().time().min)
        today_max = datetime.combine(timezone.now().date(), datetime.today().time().max)
        cantidad_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).count()
        eventos_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).values("asesor").annotate(cantidad=Count("pk"))
        cantidad_pendiente = Evento.objects.filter(fecha_hora__gt=today_max).count()
        eventos_pendientes = Evento.objects.filter(fecha_hora__gt=today_max).values("asesor").annotate(cantidad=Count("pk"))


        etapas = CatalogoRespuestasByEtapa.objects.values("etapa").distinct()
        respuestas = CatalogoRespuestasByEtapa.objects.values("respuesta").distinct()

        print(user)

        print(eventos_hoy)
        print(eventos_pendientes)

        general = True

        prospectos = Prospecto.objects.all()

        context["asesor_actual"] = asesor_actual
        context["asesores"] = asesores
        context["calendario_general"] = calendario_general
        context["cantidad_hoy"] = cantidad_hoy
        context["cantidad_pendientes"] = cantidad_pendiente
        context["etapas"] = etapas
        context["eventos"] = eventos
        context["eventos_hoy"] = eventos_hoy
        context["eventos_pendientes"] = eventos_pendientes
        context["general"] = general
        context["prospectos"] = prospectos
        context["respuestas"] = respuestas
        context["user"] = user

        return context

    def post(self, request):
        r = request.POST
        user = User.objects.get(username=self.request.user)
        
        print(r)
        if r.get("nombre_evento", None):
            nombre = r.get("nombre_evento", None)
            tipo = r.get("tipo", None)
            telefono_cliente = r.get("telefono_cliente", None)
            observaciones = r.get("observaciones", None)
            asesor = r.get("asesor", None)
            fecha_hora = r.get("fecha_hora", None)

            lead = Lead.objects.get(prospecto__celular=telefono_cliente)
            evento = Evento.objects.create(nombre=nombre,
                                           tipo=tipo,
                                           telefono_cliente=telefono_cliente,
                                           observaciones=observaciones,
                                           asesor=Asesor.objects.get(nombre=asesor),
                                           fecha_hora=datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M"),
                                           lead=lead
                                           )

            return HttpResponseRedirect(reverse_lazy('dashboards:calendar'))

        if r.get("title", None):
            nombre = r.get("title", None)
            evento = Evento.objects.get(nombre=nombre)
            evento.delete()

            return HttpResponseRedirect(reverse_lazy('dashboards:calendar'))
        if r.get("title2", None):
            nombre = r.get("title2", None)
            evento = Evento.objects.get(nombre=nombre).lead.pk
            
            print(evento)

            return JsonResponse(evento, safe=False)
        if r.get("title_cumplido", None):
            nombre = r.get("title_cumplido", None)
            evento = Evento.objects.get(nombre=nombre)
            evento.cumplido = True
            evento.fecha_hora_cumplido = datetime.now()
            evento.save()
            
            return HttpResponseRedirect(reverse_lazy('dashboards:calendar'))


class CalendarDetailView(LoginRequiredMixin, DetailView):
    # Vista de Calendar Detail

    template_name = "Calendar.html"
    slug_field = "adviser"
    slug_url_kwarg = "adviser"
    queryset = Asesor.objects.all()
    context_object_name = "adviser"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        adviser = self.get_object()
        user = User.objects.get(username=self.request.user)
        try:
            asesor_actual = Asesor.objects.get(nombre=functions.separar_nombre(user.username))
        except:
            asesor_actual = {}
            asesor_actual["pk"] = 0

        calendario_general = True
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                calendario_general = False

        asesores = Asesor.objects.all()

        eventos = Evento.objects.filter(asesor=adviser)

        today_min = datetime.combine(timezone.now().date(), datetime.today().time().min)
        today_max = datetime.combine(timezone.now().date(), datetime.today().time().max)
        cantidad_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).count()
        eventos_hoy = Evento.objects.filter(fecha_hora__range=(today_min, today_max)).values("asesor").annotate(cantidad=Count("pk"))
        cantidad_pendiente = Evento.objects.filter(fecha_hora__gt=today_max).count()
        eventos_pendientes = Evento.objects.filter(fecha_hora__gt=today_max).values("asesor").annotate(cantidad=Count("pk"))

        etapas = CatalogoRespuestasByEtapa.objects.values("etapa").distinct()
        respuestas = CatalogoRespuestasByEtapa.objects.values("respuesta").distinct()

        print(user)
        print(adviser.pk)

        mostrar_evento = False
        for grupo in self.request.user.groups.all():
            if grupo.name == "Asesor":
                mostrar_evento = True

        general = False

        prospectos = Lead.objects.filter(nombre_asesor=functions.separar_nombre(user.username)).distinct()

        print(prospectos)

        context["adviser"] = adviser
        context["asesor_actual"] = asesor_actual
        context["asesores"] = asesores
        context["calendario_general"] = calendario_general
        context["cantidad_hoy"] = cantidad_hoy
        context["cantidad_pendientes"] = cantidad_pendiente
        context["etapas"] = etapas
        context["eventos"] = eventos
        context["eventos_hoy"] = eventos_hoy
        context["eventos_pendientes"] = eventos_pendientes
        context["general"] = general
        context["mostrar_evento"] = mostrar_evento
        context["prospectos"] = prospectos
        context["respuestas"] = respuestas
        context["user"] = user

        return context

    def post(self, request, pk):
        r = request.POST
        adviser = self.get_object()
        user = User.objects.get(username=self.request.user)
        
        print(r)
        if r.get("nombre_evento", None):
            nombre = r.get("nombre_evento", None)
            tipo = r.get("tipo", None)
            telefono_cliente = r.get("telefono_cliente", None)
            observaciones = r.get("observaciones", None)
            asesor = r.get("asesor", None)
            fecha_hora = r.get("fecha_hora", None)
            
            print(telefono_cliente)
            print(asesor)
            print(type(telefono_cliente))

            lead = Lead.objects.get(prospecto__celular=telefono_cliente)
            evento = Evento.objects.create(nombre=nombre,
                                           tipo=tipo,
                                           telefono_cliente=telefono_cliente,
                                           observaciones=observaciones,
                                           asesor=Asesor.objects.get(nombre=asesor),
                                           fecha_hora=datetime.strptime(fecha_hora,"%Y-%m-%dT%H:%M"),
                                           lead=lead
                                           )

            return HttpResponseRedirect(reverse_lazy('dashboards:calendar_detail', kwargs={"pk": pk}))
        if r.get("title2", None):
            nombre = r.get("title2", None)
            telefono = r.get("telefono", None)
            try:
                Lead.objects.get(prospecto__celular=telefono, nombre_asesor=adviser.nombre)
                evento = Evento.objects.get(nombre=nombre).lead.pk
            
                return JsonResponse(evento, safe=False)
            except:
                pass

            
        if r.get("title_cumplido", None):
            nombre = r.get("title_cumplido", None)
            evento = Evento.objects.get(nombre=nombre)
            evento.cumplido = True
            evento.fecha_hora_cumplido = datetime.now()
            evento.save()
            
            return HttpResponseRedirect(reverse_lazy('dashboards:calendar_detail', kwargs={"pk": pk}))